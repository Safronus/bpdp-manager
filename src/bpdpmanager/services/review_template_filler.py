"""Heuristické vyplňování XLSX šablon posudků.

Šablony FAI UTB mají konzistentní strukturu: v sloupci A jsou popisky
(„Student:", „Vedoucí práce:", „Téma bakalářské práce:" atd.) a hodnoty
patří do sloupce B na stejném řádku. Filler tedy:

1. Otevře šablonu přes openpyxl
2. Projde sloupec A v aktivním listu, hledá popisky podle ``LABEL_PATTERNS``
3. Pro každý nalezený popisek zapíše hodnotu do sloupce B téhož řádku
4. Uloží jako *nový* XLSX (originál se nemodifikuje)

Modul je úmyslně bez závislosti na PySide6 a pydantic modelech — přijímá
slovník primitivních hodnot, aby šel snadno testovat.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

# ── Pattern → field name ────────────────────────────────────────────────────
#
# Klíče slovníku jsou regex patterny (case-insensitive) hledané v normalizované
# verzi popisku (lowercase, ASCII-folded). Hodnoty jsou klíče v ``fields``
# slovníku předaném do ``fill_template``.
#
# Filler porovnává dvě varianty popisku:
#   1. Originál (lowercase) — chytí přesnou shodu vč. diakritiky
#   2. ASCII-folded (lowercase) — pro robustnost (Téma → tema)


def _ascii_fold(s: str) -> str:
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if not unicodedata.combining(c))


# Order matters — specifické před obecnými (např. „topic of bachelor"
# musí matchnout dřív než obecné „topic:").
LABEL_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    # === Student ===
    (re.compile(r"^\s*student\s*:?\s*$", re.IGNORECASE), "student"),

    # === Vedoucí / Supervisor ===
    (re.compile(r"^\s*vedouci\s+prace\s*:?\s*$", re.IGNORECASE), "supervisor"),
    (re.compile(r"^\s*vedouci\s*:?\s*$", re.IGNORECASE), "supervisor"),
    (re.compile(r"^\s*supervisor\s+of\s+the\s+thesis\s*:?\s*$", re.IGNORECASE), "supervisor"),
    (re.compile(r"^\s*supervisor\s*:?\s*$", re.IGNORECASE), "supervisor"),

    # === Oponent / Opponent ===
    (re.compile(r"^\s*oponent\s+prace\s*:?\s*$", re.IGNORECASE), "opponent"),
    (re.compile(r"^\s*oponent\s*:?\s*$", re.IGNORECASE), "opponent"),
    (re.compile(r"^\s*opponent\s+of\s+the\s+thesis\s*:?\s*$", re.IGNORECASE), "opponent"),
    (re.compile(r"^\s*opponent\s*:?\s*$", re.IGNORECASE), "opponent"),

    # === Téma / Topic — preferujeme CZ název pro CZ šablonu, EN pro EN ===
    (re.compile(r"^\s*tema\s+bakalarske\s+prace\s*:?\s*$", re.IGNORECASE), "title_cs"),
    (re.compile(r"^\s*tema\s+diplomove\s+prace\s*:?\s*$", re.IGNORECASE), "title_cs"),
    (re.compile(r"^\s*tema\s+prace\s*:?\s*$", re.IGNORECASE), "title_cs"),
    (re.compile(r"^\s*tema\s*:?\s*$", re.IGNORECASE), "title_cs"),
    (re.compile(r"^\s*bachelor\s+thesis\s+topic\s*:?\s*$", re.IGNORECASE), "title_en"),
    (re.compile(r"^\s*master[''ʼ]?s?\s+thesis\s+topic\s*:?\s*$", re.IGNORECASE), "title_en"),
    (re.compile(r"^\s*thesis\s+topic\s*:?\s*$", re.IGNORECASE), "title_en"),
    (re.compile(r"^\s*topic\s+of\s+(bachelor|master|bp|dp)\s*:?\s*$", re.IGNORECASE), "title_en"),
    (re.compile(r"^\s*topic\s*:?\s*$", re.IGNORECASE), "title_en"),

    # === Akademický rok / Academic year ===
    (re.compile(r"^\s*akademicky\s+rok\s*:?\s*$", re.IGNORECASE), "academic_year"),
    (re.compile(r"^\s*academic\s+year\s*:?\s*$", re.IGNORECASE), "academic_year"),

    # === Studijní program / Study program (volitelné — typicky pre-filled v šabloně) ===
    (re.compile(r"^\s*studijni\s+program\s*:?\s*$", re.IGNORECASE), "study_program"),
    (re.compile(r"^\s*study\s+program\s*:?\s*$", re.IGNORECASE), "study_program"),

    # === Specializace / Specialization ===
    (re.compile(r"^\s*specializace\s*:?\s*$", re.IGNORECASE), "specialization"),
    (re.compile(r"^\s*specialization\s*:?\s*$", re.IGNORECASE), "specialization"),
]


# ── Hlavní fill funkce ──────────────────────────────────────────────────────


def fill_template(
    template_path: Path,
    output_path: Path,
    fields: dict[str, str],
    *,
    overwrite_existing_cells: bool = False,
) -> dict[str, Any]:
    """Otevře šablonu, vyplní hodnoty heuristicky podle popisků v sloupci A,
    uloží jako nový soubor.

    Args:
        template_path: Cesta k zdrojové XLSX šabloně.
        output_path: Kam zapsat vyplněnou kopii. Musí být ``.xlsx``.
        fields: Slovník hodnot. Klíče odpovídají hodnotám v ``LABEL_PATTERNS``
            (student, supervisor, opponent, title_cs, title_en,
            academic_year, study_program, specialization).
            Prázdný string nebo None pole přeskočí.
        overwrite_existing_cells: Když True, přepíše i buňky, které již
            mají hodnotu (např. defaultní „2025/2026" z šablony).
            Default False — respektujeme předvyplněné defaulty šablony,
            přepisujeme jen prázdné buňky a `academic_year` (volatile).

    Returns:
        Statistiky: ``{"filled_count": N, "skipped_existing": N,
        "skipped_unknown": N, "matches": [(coord, field, value)]}``.

    Raises:
        ImportError: openpyxl chybí.
        FileNotFoundError: template_path neexistuje.
        ValueError: cílový suffix není .xlsx.
    """
    # Lazy import openpyxl (volitelná dep — modul jen pro tuto featuru)
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "openpyxl není nainstalován — `pip install openpyxl`."
        ) from exc

    template_path = Path(template_path)
    output_path = Path(output_path)
    if not template_path.is_file():
        raise FileNotFoundError(f"Šablona neexistuje: {template_path}")
    if output_path.suffix.lower() != ".xlsx":
        raise ValueError(f"Cíl musí mít příponu .xlsx (got {output_path})")

    # Vyčisti fields — None/prázdné stringy odstraň
    clean_fields: dict[str, str] = {
        k: str(v).strip() for k, v in (fields or {}).items()
        if v is not None and str(v).strip()
    }

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb.active

    stats = {
        "filled_count": 0,
        "skipped_existing": 0,
        "skipped_unknown": 0,
        "matches": [],  # list[(coord, field, value)]
    }

    # Heuristika: walk col A, pro každý popisek najdi field
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1)
        if cell_a.value is None:
            continue
        label_raw = str(cell_a.value)
        if not label_raw.strip():
            continue

        # Match against patterns
        label_norm = _ascii_fold(label_raw).lower().strip().rstrip(":").strip()
        field_key = None
        for pattern, fkey in LABEL_PATTERNS:
            if pattern.search(label_norm) or pattern.search(label_raw.lower()):
                field_key = fkey
                break
        if field_key is None:
            continue  # popisek, který neumíme namapovat — skip

        value = clean_fields.get(field_key)
        if value is None:
            # Special case: pokud máme title_en v poli, ale šablona ptá
            # title_cs (nebo naopak), zkusíme fallback opačné jazykové
            # verze (lepší něco než nic).
            if field_key == "title_cs":
                value = clean_fields.get("title_en")
            elif field_key == "title_en":
                value = clean_fields.get("title_cs")
        if value is None:
            stats["skipped_unknown"] += 1
            continue

        # Cílová buňka je B na stejném řádku
        target = ws.cell(row=row_idx, column=2)
        # Respektuj existující default v šabloně (kromě academic_year,
        # který je vždy volatilní — chceme rok aktuální práce)
        if (
            target.value is not None
            and str(target.value).strip()
            and not overwrite_existing_cells
            and field_key != "academic_year"
        ):
            stats["skipped_existing"] += 1
            continue

        target.value = value
        stats["filled_count"] += 1
        stats["matches"].append((target.coordinate, field_key, value))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats
