"""Extrakce strukturálního schématu z XLSX šablony posudku.

Heuristika nad FAI UTB šablonami:

- Kritéria hodnocení leží v souvislé sekci, typicky řádek 17 hlavička
  (``A="Kritérium hodnocení"``, ``C="Váha"``, ``D="Body (0–5)"``)
  a následují páry (label + weight + score) / (sub-description).

- Detekce řádku jako kritéria:
    * ``A`` má neprázdný text
    * ``C`` má numerickou hodnotu (váha, např. 0.5 nebo 1.0)
    * ``D`` má numerickou hodnotu (přednastavené skóre, typicky 5)
    * ani ``C`` ani ``D`` není formule (``=...``)

- Konec sekce kritérií: řádek s formulí v ``C`` (``=SUM(...)``) nebo
  ``A=="Celkový součet vážených bodů:"``.

Pak ještě hledáme zvláštní pole pro snazší vyplňování:
- ``Splnění bodů zadání`` → buňka v ``C`` daného řádku
- ``Výsledek kontroly plagiátorství`` → ``C``
- ``Zdůvodnění výsledku kontroly plagiátorství`` → následující řádek
- ``Celkové hodnocení práce, připomínky a dotazy`` → 3 řádky dole

Modul je úmyslně bez závislosti na pydantic/PySide6 — vrací plain
dict, který volající (``ThesisService``) předá do
``ReviewTemplate.criteria`` / ``field_cells``.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any


def _ascii_fold_lower(s: str) -> str:
    if not s:
        return ""
    nfd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfd if not unicodedata.combining(c)).lower()


def _is_numeric(value: Any) -> bool:
    """True pokud value je číslo (int/float) — ne formule, ne string s číslem."""
    if isinstance(value, bool):
        return False  # bool je instance int, ale ne chceme
    if isinstance(value, (int, float)):
        return True
    return False


# Patterns pro speciální pole (CZ + EN, normalized lowercase + ASCII-fold)
_FIELD_PATTERNS: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"^\s*splneni\s+vsech\s+bodu\s+zadani"), "assignment_fulfilled"),
    (re.compile(r"^\s*fulfilment\s+of\s+all\s+assignment\s+points"), "assignment_fulfilled"),
    (re.compile(r"^\s*vysledek\s+kontroly\s+plagi"), "plagiarism_verdict"),
    (re.compile(r"^\s*plagiarism\s+check\s+result"), "plagiarism_verdict"),
    (re.compile(r"^\s*zduvodneni\s+vysledku\s+kontroly\s+plagi"), "plagiarism_justification"),
    (re.compile(r"^\s*plagiarism\s+check\s+justification"), "plagiarism_justification"),
    (re.compile(r"^\s*celkove\s+hodnoceni"), "overall_comment"),
    (re.compile(r"^\s*overall\s+evaluation"), "overall_comment"),
    (re.compile(r"^\s*misto,?\s*datum"), "place_date"),
    (re.compile(r"^\s*place,?\s*date"), "place_date"),
]


def extract_template_schema(template_path: Path) -> dict:
    """Otevře XLSX šablonu a vrátí strukturální schema.

    Returns:
        ``{
            "criteria": [
                {"row": 18, "label": "...", "default_weight": 1.0,
                 "weight_cell": "C18", "score_cell": "D18"},
                ...
            ],
            "field_cells": {
                "assignment_fulfilled": "C15",
                "plagiarism_verdict": "C39",
                "plagiarism_justification": "A41",  # nebo merge B41…
                "overall_comment": "A44",
                "place_date": "A49",
            },
        }``

    Pokud žádné kritérium nedetekuje, ``criteria`` je prázdný list — UI
    si zvolí fallback ručního vyplnění bez bodování.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl není nainstalován.") from exc

    wb = openpyxl.load_workbook(template_path, data_only=False)
    ws = wb.active

    criteria: list[dict] = []
    field_cells: dict[str, str] = {}

    max_row = min(ws.max_row, 200)  # bezpečnostní strop

    for row_idx in range(1, max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        c = ws.cell(row=row_idx, column=3).value
        d = ws.cell(row=row_idx, column=4).value

        # Kritéria — A=text, C=numeric weight, D=numeric score, žádné formule
        if (
            isinstance(a, str)
            and a.strip()
            and _is_numeric(c)
            and _is_numeric(d)
            and not (isinstance(c, str) and c.startswith("="))
            and not (isinstance(d, str) and d.startswith("="))
        ):
            criteria.append({
                "row": row_idx,
                "label": a.strip(),
                "default_weight": float(c),
                "weight_cell": f"C{row_idx}",
                "score_cell": f"D{row_idx}",
            })
            continue

        # Detekce konce sekce kritérií — formule v C (typicky =SUM nebo =C18+…)
        if isinstance(c, str) and c.startswith("=") and criteria:
            # Stačí to k tomu, abychom přestali přidávat — všechny criteria
            # už jsou před tímhle řádkem.
            pass  # noqa — nemusíme breakovat, dále hledáme field_cells

        # Speciální pole — A obsahuje label, hledáme příslušnou cell
        if isinstance(a, str) and a.strip():
            norm = _ascii_fold_lower(a)
            for pattern, key in _FIELD_PATTERNS:
                if pattern.search(norm):
                    if key in field_cells:
                        # Už máme jednou nalezeno — zachovej první match
                        continue
                    # Default: hodnota je v C na stejném řádku (assignment_fulfilled,
                    # plagiarism_verdict)
                    target_cell = f"C{row_idx}"
                    # Výjimky: justification + overall_comment + place_date jsou
                    # typicky pod labelem (víceřádkový text) → zapisujeme do
                    # následujícího řádku A.
                    if key in {"plagiarism_justification", "overall_comment"}:
                        target_cell = f"A{row_idx + 1}"
                    elif key == "place_date":
                        target_cell = f"A{row_idx}"  # přímo na řádku s "Místo, datum:"
                    field_cells[key] = target_cell
                    break

    return {"criteria": criteria, "field_cells": field_cells}
