from __future__ import annotations

import hashlib
import math
import re
import shutil
import unicodedata
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from ..config import app_data_dir, harmonograms_dir, thesis_documents_dir
from ..models import (
    AcademicYearInfo,
    Attachment,
    CriterionScore,
    KeyDate,
    Obor,
    Opponent,
    OpposingThesis,
    RejectedStudent,
    Review,
    ReviewTemplate,
    Student,
    Supervisor,
    TemplateCriterion,
    Thesis,
    ThesisProposal,
)
from ..models.enums import (
    ALLOWED_TRANSITIONS,
    AttachmentKind,
    OpponentKind,
    ThesisStatus,
    ThesisType,
)
from ..storage import Database, Repository
from .default_data import (
    DefaultTemplateSpec,
    default_obory,
    discipline_from_app_code,
    form_neutral_name,
    list_default_template_specs,
)
from .file_naming import (
    build_plagiarism_name,
    build_target_name,
    sanitize_for_fs,
    subdir_for,
)
from .harmonogram_parser import parse_pdf
from .review_pdf import extract_grade_from_file
from .review_schema import extract_template_metadata, extract_template_schema
from .review_template_filler import (
    fill_template,
    load_template_workbook,
    plan_template_fill,
)
from .xlsx_cell_writer import set_cells

# Přípony posudků, ze kterých umíme vyčíst navrženou známku (PDF + Word).
_GRADE_SOURCE_SUFFIXES = (".pdf", ".docx", ".doc")


def _is_grade_source(name: str) -> bool:
    """True, pokud jde o soubor posudku, ze kterého lze zkusit vyčíst známku."""
    return name.lower().endswith(_GRADE_SOURCE_SUFFIXES)


# Kindy příloh, kde může být víc různých souborů (dedup podle obsahu).
# Balík (text+přílohy v zipu) je taky archiv → dedupuje se obsahem jako přílohy.
_DEDUP_KINDS = {
    AttachmentKind.THESIS_APPENDIX,
    AttachmentKind.THESIS_BUNDLE,
    AttachmentKind.OTHER,
}


@dataclass
class DuplicateAppendix:
    """Jedna duplicitní příloha (stejný obsah jako ``keep_label``) k smazání."""
    work_id: str
    is_opposing: bool
    work_label: str       # student + typ práce (pro UI)
    keep_label: str       # soubor, který zůstane
    del_label: str        # duplicitní soubor ke smazání
    del_url: str          # url_or_path (identita pro smazání)
    size: int


# Přípony, které nikdy nejsou plný text práce (vždy příloha) — viz stag_api.
_ARCHIVE_EXTS = {".zip", ".rar", ".7z", ".tar", ".gz", ".tgz", ".bz2"}


@dataclass
class SwappedDocs:
    """Práce, kde je text práce a příloha prohozené (zip jako text, PDF jako
    příloha) — kandidát na nápravu. ``*_url`` jsou identity pro přeřazení."""
    work_id: str
    is_opposing: bool
    work_label: str
    text_url: str         # nyní THESIS_TEXT, ale archiv → má být příloha
    text_label: str
    appendix_url: str     # nyní příloha (PDF) → má být text práce
    appendix_label: str


@dataclass
class TextBundle:
    """Práce, kde je text práce uložený jako archiv (zip) a NENÍ k němu žádné
    samostatné PDF — jde o **balík** (text + přílohy v jednom). Kandidát na
    přeřazení z *Text práce* na *Text práce + přílohy*."""
    work_id: str
    is_opposing: bool
    work_label: str
    url: str              # nyní THESIS_TEXT (archiv) → má být THESIS_BUNDLE
    label: str


class TransitionError(ValueError):
    """Pokud se pokoušíme o nepovolený přechod mezi stavy nebo chybí povinná pole."""


class ThesisService:
    """Hlavní fasáda nad úložištěm — drží konzistenci dat a vynucuje pravidla."""

    def __init__(self, repo: Repository) -> None:
        self._repo = repo
        self._db: Database = repo.load()
        # Hloubka aktivní transakce. Když > 0, ``save()`` se odkládá až na konec
        # nejvyšší úrovně. Při výjimce uvnitř ``batch()`` se ``_db`` znovu načte
        # z disku — všechny rozdělané změny jsou zahozeny.
        self._batch_depth: int = 0

    # --- načítání / ukládání -------------------------------------------------

    @property
    def db(self) -> Database:
        return self._db

    def reload(self) -> None:
        self._db = self._repo.load()

    def reset(self, repo: Repository) -> None:
        """Nahradí podkladový repozitář a načte data z něj.

        Používá se při přepnutí na jiný profil — bez nutnosti vytvářet
        nový ThesisService a přemontovávat UI.
        """
        self._repo = repo
        self._db = repo.load()

    def save(self) -> None:
        """Persist ``_db`` na disk. V rámci ``batch()`` se zápis odkládá."""
        if self._batch_depth > 0:
            return  # odložit; reálné save proběhne na konci batche
        self._repo.save(self._db)

    @contextmanager
    def batch(self) -> Iterator[None]:
        """Transakční blok — všechny ``upsert_*`` / ``delete_*`` uvnitř
        nepíšou na disk. Save proběhne až na konci, atomicky.

        Při výjimce uvnitř se ``_db`` znovu načte z disku → všechny změny
        provedené v rámci tohoto bloku se zahodí (rollback).

        Pozor: copy souborů (``attach_document`` apod.) je *durable* —
        soubory zůstanou na disku i po rollbacku, ale bez záznamu v ``_db``
        se z hlediska aplikace stávají neviditelnými („orphan files").
        Mazání orphanu je out-of-scope této transakce.
        """
        self._batch_depth += 1
        try:
            yield
        except Exception:
            self._batch_depth -= 1
            if self._batch_depth == 0:
                # Zahoď in-memory změny, vrať na poslední uložený stav
                self._db = self._repo.load()
            raise
        else:
            self._batch_depth -= 1
            if self._batch_depth == 0:
                self._repo.save(self._db)

    # --- studenti ------------------------------------------------------------

    def list_students(self) -> list[Student]:
        return sorted(self._db.students, key=lambda s: (s.last_name.lower(), s.first_name.lower()))

    def get_student(self, student_id: str) -> Student | None:
        return next((s for s in self._db.students if s.id == student_id), None)

    def upsert_student(self, student: Student) -> Student:
        existing = self.get_student(student.id)
        if existing:
            idx = self._db.students.index(existing)
            self._db.students[idx] = student
        else:
            self._db.students.append(student)
        self.save()
        return student

    def delete_student(self, student_id: str) -> None:
        self._db.students = [s for s in self._db.students if s.id != student_id]
        for t in self._db.theses:
            if t.student_id == student_id:
                t.student_id = None
                t.touch()
        self.save()

    # --- odmítnutí zájemci ---------------------------------------------------

    def list_rejected_students(self) -> list[RejectedStudent]:
        return sorted(
            self._db.rejected_students,
            key=lambda r: (r.academic_year, r.name.lower()),
            reverse=True,
        )

    def upsert_rejected_student(self, rej: RejectedStudent) -> RejectedStudent:
        existing = next(
            (r for r in self._db.rejected_students if r.id == rej.id), None
        )
        if existing:
            self._db.rejected_students[self._db.rejected_students.index(existing)] = rej
        else:
            self._db.rejected_students.append(rej)
        self.save()
        return rej

    def remove_rejected_student(self, rej_id: str) -> None:
        self._db.rejected_students = [
            r for r in self._db.rejected_students if r.id != rej_id
        ]
        self.save()

    # --- oponenti ------------------------------------------------------------

    def list_opponents(self, kind: OpponentKind | None = None) -> list[Opponent]:
        opps = self._db.opponents
        if kind is not None:
            opps = [o for o in opps if o.kind == kind]
        return sorted(opps, key=lambda o: o.name.lower())

    def get_opponent(self, opponent_id: str) -> Opponent | None:
        return next((o for o in self._db.opponents if o.id == opponent_id), None)

    def upsert_opponent(self, opponent: Opponent) -> Opponent:
        existing = self.get_opponent(opponent.id)
        if existing:
            idx = self._db.opponents.index(existing)
            self._db.opponents[idx] = opponent
        else:
            self._db.opponents.append(opponent)
        self.save()
        return opponent

    def delete_opponent(self, opponent_id: str) -> None:
        self._db.opponents = [o for o in self._db.opponents if o.id != opponent_id]
        for t in self._db.theses:
            if t.opponent_id == opponent_id:
                t.opponent_id = None
                t.touch()
        self.save()

    # --- vedoucí (pro oponentské posudky) -----------------------------------

    def list_supervisors(self) -> list[Supervisor]:
        return sorted(self._db.supervisors, key=lambda s: s.name.lower())

    def get_supervisor(self, supervisor_id: str) -> Supervisor | None:
        return next(
            (s for s in self._db.supervisors if s.id == supervisor_id), None
        )

    def get_supervisor_by_name(self, name: str) -> Supervisor | None:
        """Najde vedoucího podle přesné shody jména (case-sensitive)."""
        if not name:
            return None
        return next(
            (s for s in self._db.supervisors if s.name == name), None
        )

    def upsert_supervisor(self, supervisor: Supervisor) -> Supervisor:
        existing = self.get_supervisor(supervisor.id)
        if existing:
            idx = self._db.supervisors.index(existing)
            self._db.supervisors[idx] = supervisor
        else:
            self._db.supervisors.append(supervisor)
        self.save()
        return supervisor

    def delete_supervisor(self, supervisor_id: str) -> None:
        """Smaže vedoucího z registry. Inline údaje v oponentských posudcích
        zůstávají (jsou kopií, ne FK)."""
        self._db.supervisors = [
            s for s in self._db.supervisors if s.id != supervisor_id
        ]
        self.save()

    # --- úklid jmen s tituly (parsing „Příjmení Jméno, tituly" ze STAG) -------

    def cleanup_opponent_titles(self, *, dry_run: bool = False) -> list[tuple[str, str]]:
        """Rozparsuje jména oponentů s tituly do polí (title_before/name/after).

        Týká se záznamů, jejichž ``name`` obsahuje čárku (formát ze STAG). Vrací
        seznam ``(staré zobrazení, nové zobrazení)``; při ``dry_run`` neukládá.
        """
        from ..models.naming import compose_titled_name, parse_titled_name

        changes: list[tuple[str, str]] = []
        for o in self._db.opponents:
            if "," not in (o.name or ""):
                continue
            before, name, after = parse_titled_name(o.name)
            if not name or (before, name, after) == (
                o.title_before, o.name, o.title_after
            ):
                continue
            old = compose_titled_name(o.title_before, o.name, o.title_after)
            changes.append((old, compose_titled_name(before, name, after)))
            if not dry_run:
                o.title_before, o.name, o.title_after = before, name, after
        if changes and not dry_run:
            self.save()
        return changes

    def cleanup_supervisor_titles(self, *, dry_run: bool = False) -> list[tuple[str, str]]:
        """Rozparsuje jména vedoucích s tituly + srovná denormalizované
        ``OpposingThesis.supervisor_name`` (kopie stringu, ne FK)."""
        from ..models.naming import compose_titled_name, parse_titled_name

        changes: list[tuple[str, str]] = []
        for s in self._db.supervisors:
            if "," not in (s.name or ""):
                continue
            before, name, after = parse_titled_name(s.name)
            if not name or (before, name, after) == (
                s.title_before, s.name, s.title_after
            ):
                continue
            old = compose_titled_name(s.title_before, s.name, s.title_after)
            changes.append((old, compose_titled_name(before, name, after)))
            if not dry_run:
                s.title_before, s.name, s.title_after = before, name, after
        # Denormalizované jméno vedoucího u oponentur (jen string, přeskládáme).
        for op in self._db.opposing_theses:
            raw = op.supervisor_name or ""
            if "," not in raw:
                continue
            before, name, after = parse_titled_name(raw)
            new = compose_titled_name(before, name, after)
            if new and new != raw:
                changes.append((raw, new))
                if not dry_run:
                    op.supervisor_name = new
        if changes and not dry_run:
            self.save()
        return changes

    # --- obory ---------------------------------------------------------------

    def list_obory(self) -> list[str]:
        """Vrátí seznam názvů oborů — pro plnění combo boxů (backward compat)."""
        return sorted(o.name for o in self._db.obory)

    def list_obor_objects(self) -> list[Obor]:
        """Vrátí seznam Obor objektů (včetně kontaktu na sekretářku)."""
        return sorted(self._db.obory, key=lambda o: o.name.lower())

    def get_obor(self, name: str) -> Obor | None:
        return next((o for o in self._db.obory if o.name == name), None)

    def get_obor_by_stag_code(self, stag_code: str) -> Obor | None:
        if not stag_code:
            return None
        return next(
            (o for o in self._db.obory if (o.stag_code or "") == stag_code), None
        )

    def add_obor(self, name: str) -> Obor | None:
        """Přidá obor (pokud neexistuje). Vrací příslušný Obor."""
        name = name.strip()
        if not name:
            return None
        existing = self.get_obor(name)
        if existing:
            return existing
        obor = Obor(name=name)
        self._db.obory.append(obor)
        self.save()
        return obor

    def upsert_obor(self, obor: Obor) -> Obor:
        """Vloží nebo aktualizuje obor (klíč = ``name``)."""
        existing = self.get_obor(obor.name)
        if existing:
            idx = self._db.obory.index(existing)
            self._db.obory[idx] = obor
        else:
            self._db.obory.append(obor)
        self.save()
        return obor

    def rename_obor(self, old_name: str, new_name: str) -> int:
        """Přejmenuje obor v číselníku a u všech studentů s tímto oborem.

        Vrací počet aktualizovaných studentů.
        """
        new_name = new_name.strip()
        if not new_name or old_name == new_name:
            return 0
        old_obor = self.get_obor(old_name)
        target = self.get_obor(new_name)
        if old_obor is not None:
            if target is not None and target is not old_obor:
                # cíl už existuje — sloučení: starý zahodíme, ponecháme target
                self._db.obory = [o for o in self._db.obory if o is not old_obor]
            else:
                old_obor.name = new_name
        count = 0
        for student in self._db.students:
            if student.obor == old_name:
                student.obor = new_name
                count += 1
        self.save()
        return count

    def remove_obor(self, name: str) -> int:
        """Smaže obor z číselníku. Studentům s tímto oborem ho vyprázdní.

        Vrací počet ovlivněných studentů.
        """
        self._db.obory = [o for o in self._db.obory if o.name != name]
        count = 0
        for student in self._db.students:
            if student.obor == name:
                student.obor = ""
                count += 1
        self.save()
        return count

    def obor_usage_count(self, name: str) -> int:
        return sum(1 for s in self._db.students if s.obor == name)

    # --- akademické roky -----------------------------------------------------

    def list_academic_years(self) -> list[str]:
        years = {t.academic_year for t in self._db.theses if t.academic_year}
        return sorted(years, reverse=True)

    @staticmethod
    def current_academic_year() -> str:
        today = date.today()
        start = today.year if today.month >= 9 else today.year - 1
        return f"{start}/{start + 1}"

    @staticmethod
    def next_academic_year() -> str:
        cur = ThesisService.current_academic_year()
        start = int(cur.split("/")[0]) + 1
        return f"{start}/{start + 1}"

    @staticmethod
    def previous_academic_year() -> str:
        cur = ThesisService.current_academic_year()
        start = int(cur.split("/")[0]) - 1
        return f"{start}/{start + 1}"

    # --- práce ---------------------------------------------------------------

    def list_theses(self) -> list[Thesis]:
        return list(self._db.theses)

    def get_thesis(self, thesis_id: str) -> Thesis | None:
        return next((t for t in self._db.theses if t.id == thesis_id), None)

    def upsert_thesis(self, thesis: Thesis) -> Thesis:
        thesis.touch()
        existing = self.get_thesis(thesis.id)
        if existing:
            idx = self._db.theses.index(existing)
            self._db.theses[idx] = thesis
        else:
            self._db.theses.append(thesis)
        self.save()
        return thesis

    def delete_thesis(self, thesis_id: str) -> None:
        self._db.theses = [t for t in self._db.theses if t.id != thesis_id]
        self.save()

    def reclassify_defense_records(
        self, *, dry_run: bool = False
    ) -> list[tuple[str, str]]:
        """Přeřadí přílohy typu „Jiné" vypadající jako protokol/zápis o průběhu
        obhajoby na typ ``DEFENSE_RECORD``.

        Vrací seznam ``(popis práce, název souboru)`` přeřazených příloh. Při
        ``dry_run=True`` jen vrátí kandidáty bez zápisu (pro náhled/potvrzení).
        """
        from .stag_api import is_defense_record_filename

        def _match(a) -> bool:
            return (
                a.kind == AttachmentKind.OTHER
                and a.is_file
                and is_defense_record_filename(a.label or a.url_or_path)
            )

        changes: list[tuple[str, str]] = []
        for t in self._db.theses:
            student = self.get_student(t.student_id) if t.student_id else None
            name = student.full_name if student else "(bez studenta)"
            label = f"{name} — {t.type.value} {t.academic_year}".strip()
            for a in t.attachments:
                if _match(a):
                    changes.append((label, a.label))
                    if not dry_run:
                        a.kind = AttachmentKind.DEFENSE_RECORD
        for o in self._db.opposing_theses:
            name = f"{o.student_last_name} {o.student_first_name}".strip() or "(student)"
            label = f"{name} — {o.type.value} {o.academic_year} (oponentura)".strip()
            for a in o.attachments:
                if _match(a):
                    changes.append((label, a.label))
                    if not dry_run:
                        a.kind = AttachmentKind.DEFENSE_RECORD
        if changes and not dry_run:
            self.save()
        return changes

    # --- návrhy témat (proposals) -------------------------------------------

    def list_proposals(self) -> list[ThesisProposal]:
        return list(self._db.proposals)

    def get_proposal(self, proposal_id: str) -> ThesisProposal | None:
        return next((p for p in self._db.proposals if p.id == proposal_id), None)

    def upsert_proposal(self, proposal: ThesisProposal) -> ThesisProposal:
        proposal.touch()
        existing = self.get_proposal(proposal.id)
        if existing:
            idx = self._db.proposals.index(existing)
            self._db.proposals[idx] = proposal
        else:
            self._db.proposals.append(proposal)
        self.save()
        return proposal

    def delete_proposal(self, proposal_id: str) -> None:
        self._db.proposals = [p for p in self._db.proposals if p.id != proposal_id]
        self.save()

    def convert_proposal_to_thesis(self, proposal_id: str) -> Thesis:
        """Z návrhu tématu založí novou vedenou práci a návrh odebere.

        Přenese typ (BP/DP), název, popis (→ anotace), body zadání a literaturu.
        Stav = *Zájemce s tématem* (entry), akademický rok = aktuální. Obor se
        nepřenáší (drží ho až student, kterého přiřadíš). Vrací novou práci.
        """
        p = self.get_proposal(proposal_id)
        if p is None:
            raise ValueError(f"Návrh {proposal_id} neexistuje.")
        thesis = Thesis(
            type=p.type,
            status=ThesisStatus.RESERVED,
            academic_year=self.current_academic_year(),
            title_cs=p.title_cs,
            annotation=p.description,
            objectives=p.objectives,
            references=p.references,
        )
        self._db.theses.append(thesis)
        self._db.proposals = [x for x in self._db.proposals if x.id != proposal_id]
        self.save()
        return thesis

    def rollback_thesis(self, thesis_id: str) -> dict[str, int]:
        """Kompletně smaže práci z DB **i** všechny její soubory na disku.

        Použití: oprava chybného importu, omyl při zakládání, ...
        Studenta/oponenta v registru nemažeme — mohou být provázáni s jinými
        pracemi a tady nemáme dost kontextu na rozhodnutí.

        Vrací statistiky: ``{"files_deleted": N, "plagiarism_pdf": 0|1,
        "documents_dir_removed": 0|1}``.
        """
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return {"files_deleted": 0, "plagiarism_pdf": 0, "documents_dir_removed": 0}

        stats = {"files_deleted": 0, "plagiarism_pdf": 0, "documents_dir_removed": 0}
        docs_dir = thesis_documents_dir(thesis_id)

        # 1) Smaž každý záznam přílohy explicitně (kvůli statistice). Adresář
        #    poté smažeme rekurzivně — chytí i orphan soubory bez záznamu.
        for att in list(thesis.attachments):
            try:
                p = docs_dir / att.url_or_path
                if p.is_file():
                    p.unlink()
                    stats["files_deleted"] += 1
            except OSError:
                pass

        # 2) Plagiátorský PDF protokol
        if thesis.plagiarism_pdf_filename:
            try:
                p = docs_dir / thesis.plagiarism_pdf_filename
                if p.is_file():
                    p.unlink()
                    stats["plagiarism_pdf"] = 1
            except OSError:
                pass

        # 3) Smaž celou složku ``documents/{id}/`` rekurzivně (chytí orphany)
        try:
            if docs_dir.exists():
                shutil.rmtree(docs_dir, ignore_errors=True)
                stats["documents_dir_removed"] = 1
        except OSError:
            pass

        # 4) Vyřaď z DB
        self._db.theses = [t for t in self._db.theses if t.id != thesis_id]
        self.save()
        return stats

    def rollback_opposing_thesis(self, op_id: str) -> dict[str, int]:
        """Analog ``rollback_thesis`` pro oponentský posudek."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return {"files_deleted": 0, "documents_dir_removed": 0}

        stats = {"files_deleted": 0, "documents_dir_removed": 0}
        docs_dir = thesis_documents_dir(f"opposing-{op_id}")

        for att in list(op.attachments):
            try:
                p = docs_dir / att.url_or_path
                if p.is_file():
                    p.unlink()
                    stats["files_deleted"] += 1
            except OSError:
                pass

        try:
            if docs_dir.exists():
                shutil.rmtree(docs_dir, ignore_errors=True)
                stats["documents_dir_removed"] = 1
        except OSError:
            pass

        self._db.opposing_theses = [
            t for t in self._db.opposing_theses if t.id != op_id
        ]
        self.save()
        return stats

    def rollback_preview(self, thesis_id: str) -> dict:
        """Spočítá, co by ``rollback_thesis`` smazalo — pro confirmation dialog.

        Vrací: ``{"thesis": Thesis, "attachments": [(label, path, exists)],
        "plagiarism_pdf": (filename, exists) | None, "extra_files": [path],
        "total_bytes": int}``.
        """
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return {"thesis": None, "attachments": [], "plagiarism_pdf": None,
                    "extra_files": [], "total_bytes": 0}

        docs_dir = thesis_documents_dir(thesis_id)
        atts: list[tuple[str, Path, bool, int]] = []
        tracked_paths: set[Path] = set()
        for att in thesis.attachments:
            p = docs_dir / att.url_or_path
            size = p.stat().st_size if p.is_file() else 0
            atts.append((att.label or att.url_or_path, p, p.is_file(), size))
            tracked_paths.add(p.resolve())

        plagiarism = None
        if thesis.plagiarism_pdf_filename:
            p = docs_dir / thesis.plagiarism_pdf_filename
            plagiarism = (thesis.plagiarism_pdf_filename, p, p.is_file(),
                          p.stat().st_size if p.is_file() else 0)
            tracked_paths.add(p.resolve())

        # Orphan soubory ve složce — vše ostatní co tam je
        extra: list[tuple[Path, int]] = []
        if docs_dir.exists():
            for p in docs_dir.rglob("*"):
                if p.is_file() and p.resolve() not in tracked_paths:
                    extra.append((p, p.stat().st_size))

        total = (
            sum(s for _, _, _, s in atts)
            + (plagiarism[3] if plagiarism else 0)
            + sum(s for _, s in extra)
        )
        return {
            "thesis": thesis,
            "attachments": atts,
            "plagiarism_pdf": plagiarism,
            "extra_files": extra,
            "total_bytes": total,
            "documents_dir": docs_dir,
        }

    def rollback_opposing_preview(self, op_id: str) -> dict:
        """Preview pro ``rollback_opposing_thesis``."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return {"opposing": None, "attachments": [], "extra_files": [], "total_bytes": 0}

        docs_dir = thesis_documents_dir(f"opposing-{op_id}")
        atts: list[tuple[str, Path, bool, int]] = []
        tracked_paths: set[Path] = set()
        for att in op.attachments:
            p = docs_dir / att.url_or_path
            size = p.stat().st_size if p.is_file() else 0
            atts.append((att.label or att.url_or_path, p, p.is_file(), size))
            tracked_paths.add(p.resolve())

        extra: list[tuple[Path, int]] = []
        if docs_dir.exists():
            for p in docs_dir.rglob("*"):
                if p.is_file() and p.resolve() not in tracked_paths:
                    extra.append((p, p.stat().st_size))

        total = sum(s for _, _, _, s in atts) + sum(s for _, s in extra)
        return {
            "opposing": op,
            "attachments": atts,
            "extra_files": extra,
            "total_bytes": total,
            "documents_dir": docs_dir,
        }

    def transition(self, thesis_id: str, target: ThesisStatus) -> Thesis:
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            raise TransitionError(f"Práce {thesis_id} neexistuje.")

        if thesis.status == target:
            return thesis

        allowed = ALLOWED_TRANSITIONS.get(thesis.status, set())
        if target not in allowed:
            raise TransitionError(
                f"Přechod {thesis.status.label} → {target.label} není povolen."
            )

        if target == ThesisStatus.LISTED:
            ok, missing = thesis.is_ready_for_listing()
            if not ok:
                raise TransitionError(
                    f"Pro vypsání tématu chybí: {', '.join(missing)}."
                )
        if target == ThesisStatus.IN_PROGRESS:
            # Vstup do V řešení (= dřívější ASSIGNED + IN_PROGRESS) vyžaduje
            # úplné oficiální zadání (titul EN, body zadání, literatura).
            # Výjimka: druhý pokus obhajoby (Nedokončeno / Neobhájeno →
            # IN_PROGRESS) — tam už zadání jednou bylo, neblokujeme.
            if thesis.status not in (ThesisStatus.CANCELLED, ThesisStatus.FAILED):
                ok, missing = thesis.is_ready_for_assignment()
                if not ok:
                    raise TransitionError(
                        f"Pro spuštění práce chybí: {', '.join(missing)}."
                    )

        thesis.status = target
        thesis.touch()
        self.save()
        return thesis

    # --- pohledy -------------------------------------------------------------

    def theses_by_year_and_type(self) -> dict[str, dict[str, list[Thesis]]]:
        out: dict[str, dict[str, list[Thesis]]] = {}
        for t in self._db.theses:
            year = t.academic_year or "(bez roku)"
            out.setdefault(year, {"BP": [], "DP": []})
            out[year][t.type.value].append(t)
        return out

    # --- harmonogram ---------------------------------------------------------

    def get_year_info(self, label: str) -> AcademicYearInfo | None:
        return next((y for y in self._db.academic_years if y.label == label), None)

    def get_or_create_year_info(self, label: str) -> AcademicYearInfo:
        info = self.get_year_info(label)
        if info is None:
            info = AcademicYearInfo(label=label)
            self._db.academic_years.append(info)
            self.save()
        return info

    def list_year_infos(self) -> list[AcademicYearInfo]:
        return sorted(self._db.academic_years, key=lambda y: y.label, reverse=True)

    def upsert_year_info(self, info: AcademicYearInfo) -> AcademicYearInfo:
        existing = self.get_year_info(info.label)
        if existing:
            idx = self._db.academic_years.index(existing)
            self._db.academic_years[idx] = info
        else:
            self._db.academic_years.append(info)
        self.save()
        return info

    def import_harmonogram_pdf(self, label: str, source_pdf: Path) -> AcademicYearInfo:
        """Importuje PDF harmonogramu pro daný akademický rok.

        - Zkopíruje PDF do ``~/.bpdpmanager/harmonograms/{label}.pdf``
        - Spustí parser a uloží extrahované klíčové termíny
        """
        info = self.get_or_create_year_info(label)
        target_name = label.replace("/", "-") + ".pdf"
        target_path = harmonograms_dir() / target_name
        shutil.copy2(source_pdf, target_path)

        info.pdf_filename = target_name
        info.pdf_source_path = str(source_pdf)

        parsed = parse_pdf(target_path)
        # zachovej manuálně přidané, nahraď jen ty importované
        manual = [kd for kd in info.key_dates if kd.source == "manual"]
        info.key_dates = manual + parsed

        return self.upsert_year_info(info)

    def add_key_date(self, year_label: str, key_date: KeyDate) -> AcademicYearInfo:
        info = self.get_or_create_year_info(year_label)
        info.key_dates.append(key_date)
        return self.upsert_year_info(info)

    def update_key_date(self, year_label: str, index: int, key_date: KeyDate) -> AcademicYearInfo:
        info = self.get_or_create_year_info(year_label)
        if 0 <= index < len(info.key_dates):
            info.key_dates[index] = key_date
        return self.upsert_year_info(info)

    def remove_key_date(self, year_label: str, index: int) -> AcademicYearInfo:
        info = self.get_or_create_year_info(year_label)
        if 0 <= index < len(info.key_dates):
            del info.key_dates[index]
        return self.upsert_year_info(info)

    # --- dokumenty k práci ---------------------------------------------------

    @staticmethod
    def _version_and_supersede(
        attachments: list[Attachment], kind: AttachmentKind, label: str
    ) -> int:
        """Vrátí novou verzi a přepne stávající ``current`` na False.

        U **víceinstančních** kindů (přílohy, OTHER) může být víc různých
        souborů zároveň aktuálních — verzuje a přepíná se proto jen v rámci
        **stejného názvu** (``label``). Jinak by druhá příloha (např.
        ``…_part2.zip``) nahradila první (``…_part1.zip``) jako „starou verzi".
        U jednoinstančních kindů (text práce, posudky, STAG export) se verzuje
        proti celému kindu jako dosud.
        """
        if kind in _DEDUP_KINDS:
            same = [a for a in attachments if a.kind == kind and a.label == label]
        else:
            same = [a for a in attachments if a.kind == kind]
        next_version = max((a.version for a in same), default=0) + 1
        for a in same:
            a.is_current = False
        return next_version

    @staticmethod
    def _file_sha256(path: Path) -> str | None:
        try:
            h = hashlib.sha256()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(1 << 20), b""):
                    h.update(chunk)
            return h.hexdigest()
        except OSError:
            return None

    def _abs_attachment_path(
        self, work_id: str, att: Attachment, *, opposing: bool
    ) -> Path | None:
        return (
            self.opposing_document_absolute_path(work_id, att) if opposing
            else self.document_absolute_path(work_id, att)
        )

    def _find_identical_attachment(
        self, work_id: str, attachments: list[Attachment],
        source_path: Path, *, opposing: bool, kind: AttachmentKind,
    ) -> Attachment | None:
        """Najde už přiloženou přílohu se **shodným obsahem** (velikost + hash),
        aby se stejný soubor nepřidával duplicitně. Hashuje jen při shodě velikosti."""
        try:
            size = source_path.stat().st_size
        except OSError:
            return None
        src_hash: str | None = None
        for a in attachments:
            if not a.is_file or a.kind != kind:
                continue
            p = self._abs_attachment_path(work_id, a, opposing=opposing)
            if p is None or not p.is_file() or p.stat().st_size != size:
                continue
            if src_hash is None:
                src_hash = self._file_sha256(source_path)
                if src_hash is None:
                    return None
            if self._file_sha256(p) == src_hash:
                return a
        return None

    def _work_label(self, work, is_opposing: bool) -> str:
        if is_opposing:
            name = work.student_full_name or "(neuvedený student)"
            return f"{name} · {work.type.value} (oponent)"
        student = self.get_student(work.student_id) if work.student_id else None
        name = student.full_name if student else (work.display_title or "—")
        return f"{name} · {work.type.value}"

    def find_duplicate_appendices(self) -> list[DuplicateAppendix]:
        """Najde přílohy (a *Jiné*) se **shodným obsahem** v rámci jedné práce
        — kandidáty na úklid. Posudky/text práce se neřeší."""
        out: list[DuplicateAppendix] = []

        def scan(works, is_opposing: bool) -> None:
            for w in works:
                groups: dict[tuple, list[Attachment]] = {}
                for a in w.attachments:
                    if not a.is_file or a.kind not in _DEDUP_KINDS:
                        continue
                    p = self._abs_attachment_path(w.id, a, opposing=is_opposing)
                    if p is None or not p.is_file():
                        continue
                    h = self._file_sha256(p)
                    if h is None:
                        continue
                    groups.setdefault((p.stat().st_size, h), []).append(a)
                for (size, _h), group in groups.items():
                    if len(group) < 2:
                        continue
                    ordered = sorted(
                        group, key=lambda a: (0 if a.is_current else 1, -a.version)
                    )
                    keep = ordered[0]
                    for dup in ordered[1:]:
                        out.append(DuplicateAppendix(
                            work_id=w.id, is_opposing=is_opposing,
                            work_label=self._work_label(w, is_opposing),
                            keep_label=keep.label, del_label=dup.label,
                            del_url=dup.url_or_path, size=size,
                        ))

        scan(self.list_theses(), False)
        scan(self.list_opposing_theses(), True)
        return out

    def delete_appendix_duplicates(
        self, items: list[tuple[str, bool, str]]
    ) -> int:
        """Smaže duplicitní přílohy (soubor + evidence). ``items`` =
        ``(work_id, is_opposing, del_url)``. Zbylé přílohy práce se označí jako
        aktuální (žádná omylem „archivovaná"). Vrací počet smazaných."""
        from collections import defaultdict

        by_work: dict[tuple[str, bool], set[str]] = defaultdict(set)
        for work_id, is_opposing, del_url in items:
            by_work[(work_id, is_opposing)].add(del_url)

        removed = 0
        for (work_id, is_opposing), urls in by_work.items():
            work = (self.get_opposing_thesis(work_id) if is_opposing
                    else self.get_thesis(work_id))
            if work is None:
                continue
            kept: list[Attachment] = []
            for a in work.attachments:
                if a.is_file and a.url_or_path in urls:
                    p = self._abs_attachment_path(work_id, a, opposing=is_opposing)
                    try:
                        if p and p.is_file():
                            p.unlink()
                    except OSError:
                        pass
                    removed += 1
                else:
                    kept.append(a)
            work.attachments = kept
            for a in work.attachments:
                if a.kind in _DEDUP_KINDS:
                    a.is_current = True
            if is_opposing:
                self.upsert_opposing_thesis(work)
            else:
                self.upsert_thesis(work)
        return removed

    # --- náprava prohozeného textu/přílohy (STAG „el. podoba" dle pořadí) -----

    @staticmethod
    def _is_archive(url_or_path: str) -> bool:
        return Path(url_or_path).suffix.lower() in _ARCHIVE_EXTS

    def find_swapped_documents(self) -> list[SwappedDocs]:
        """Najde práce, kde je **text práce archiv** (zip…) a zároveň je k dispozici
        **PDF příloha** — typický příznak, že STAG soubory přišly v jiném pořadí a
        starší detekce je prohodila. Vrací jen **jednoznačné** páry (právě jeden
        archiv-jako-text a právě jedno PDF mezi přílohami); nejasné případy
        (víc kandidátů) přeskočí — ty se řeší ručně."""
        out: list[SwappedDocs] = []

        def scan(works, is_opposing: bool) -> None:
            for w in works:
                files = [a for a in w.attachments if a.is_file]
                text_archives = [
                    a for a in files
                    if a.kind == AttachmentKind.THESIS_TEXT
                    and self._is_archive(a.url_or_path)
                ]
                pdf_appendices = [
                    a for a in files
                    if a.kind in _DEDUP_KINDS
                    and Path(a.url_or_path).suffix.lower() == ".pdf"
                ]
                # Jednoznačný prohoz: přesně jeden archiv-text a přesně jedno PDF.
                if len(text_archives) == 1 and len(pdf_appendices) == 1:
                    ta, pa = text_archives[0], pdf_appendices[0]
                    out.append(SwappedDocs(
                        work_id=w.id, is_opposing=is_opposing,
                        work_label=self._work_label(w, is_opposing),
                        text_url=ta.url_or_path, text_label=ta.label,
                        appendix_url=pa.url_or_path, appendix_label=pa.label,
                    ))

        scan(self.list_theses(), False)
        scan(self.list_opposing_theses(), True)
        return out

    def _reclassify_file_attachment(
        self, work, att: Attachment, new_kind: AttachmentKind, *, opposing: bool
    ) -> None:
        """Přesune soubor přílohy do správné podsložky, přejmenuje podle nového
        druhu a přepíše ``kind``/``label``/``url_or_path``. Nemění obsah souboru."""
        old_abs = self._abs_attachment_path(work.id, att, opposing=opposing)
        if old_abs is None or not old_abs.is_file():
            return
        surname = (work.student_last_name if opposing
                   else self._student_surname_for_thesis(work))
        base_id = f"opposing-{work.id}" if opposing else work.id
        subdir = subdir_for(new_kind)
        target_dir = thesis_documents_dir(base_id) / subdir
        target_dir.mkdir(parents=True, exist_ok=True)
        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        new_name = build_target_name(
            surname, new_kind, old_abs, existing_names=existing,
            orig_name=att.label,
        )
        target_path = target_dir / new_name
        shutil.move(str(old_abs), str(target_path))
        att.kind = new_kind
        att.label = new_name
        att.url_or_path = f"{subdir}/{new_name}"
        att.is_current = True

    def repair_swapped_documents(
        self, items: list[tuple[str, bool, str, str]]
    ) -> int:
        """Prohodí druh text práce ↔ příloha. ``items`` =
        ``(work_id, is_opposing, text_url, appendix_url)`` — ``text_url`` je
        archiv (stane se přílohou), ``appendix_url`` je PDF (stane se textem).
        Soubory se zároveň přejmenují a přesunou do správné podsložky. Před
        zápisem si volající zajistí zálohu. Vrací počet opravených prací."""
        repaired = 0
        for work_id, is_opposing, text_url, appendix_url in items:
            work = (self.get_opposing_thesis(work_id) if is_opposing
                    else self.get_thesis(work_id))
            if work is None:
                continue
            archive_att = next(
                (a for a in work.attachments if a.url_or_path == text_url), None
            )
            pdf_att = next(
                (a for a in work.attachments if a.url_or_path == appendix_url), None
            )
            if archive_att is None or pdf_att is None:
                continue
            # Nejdřív PDF → text práce, pak archiv → příloha.
            self._reclassify_file_attachment(
                work, pdf_att, AttachmentKind.THESIS_TEXT, opposing=is_opposing
            )
            self._reclassify_file_attachment(
                work, archive_att, AttachmentKind.THESIS_APPENDIX,
                opposing=is_opposing,
            )
            if is_opposing:
                self.upsert_opposing_thesis(work)
            else:
                self.upsert_thesis(work)
            repaired += 1
        return repaired

    def find_text_bundles(self) -> list[TextBundle]:
        """Najde práce, kde je **text práce archiv** (zip…) a zároveň u práce
        **NENÍ žádné PDF mezi přílohami** — jde o **balík** (text + přílohy
        v jednom zipu), ne o prohození. Kandidáti na přeřazení z *Text práce*
        na *Text práce + přílohy*. (Práce s archivem-textem i PDF-přílohou řeší
        ``find_swapped_documents`` jako prohození.)"""
        out: list[TextBundle] = []

        def scan(works, is_opposing: bool) -> None:
            for w in works:
                files = [a for a in w.attachments if a.is_file]
                text_archives = [
                    a for a in files
                    if a.kind == AttachmentKind.THESIS_TEXT
                    and self._is_archive(a.url_or_path)
                ]
                pdf_appendices = [
                    a for a in files
                    if a.kind in _DEDUP_KINDS
                    and Path(a.url_or_path).suffix.lower() == ".pdf"
                ]
                # Balík = archiv jako text BEZ jakéhokoli PDF mezi přílohami.
                if text_archives and not pdf_appendices:
                    for a in text_archives:
                        out.append(TextBundle(
                            work_id=w.id, is_opposing=is_opposing,
                            work_label=self._work_label(w, is_opposing),
                            url=a.url_or_path, label=a.label,
                        ))

        scan(self.list_theses(), False)
        scan(self.list_opposing_theses(), True)
        return out

    def reclassify_text_bundles(
        self, items: list[tuple[str, bool, str]]
    ) -> int:
        """Přeřadí archiv vedený jako *Text práce* na *Text práce + přílohy*.
        ``items`` = ``(work_id, is_opposing, url)``. Soubor se přejmenuje/přesune
        do podsložky balíku. Před zápisem si volající zajistí zálohu. Vrací
        počet přeřazených souborů."""
        from collections import defaultdict

        by_work: dict[tuple[str, bool], set[str]] = defaultdict(set)
        for work_id, is_opposing, url in items:
            by_work[(work_id, is_opposing)].add(url)

        changed = 0
        for (work_id, is_opposing), urls in by_work.items():
            work = (self.get_opposing_thesis(work_id) if is_opposing
                    else self.get_thesis(work_id))
            if work is None:
                continue
            touched = False
            for att in list(work.attachments):
                if att.url_or_path in urls and att.kind == AttachmentKind.THESIS_TEXT:
                    self._reclassify_file_attachment(
                        work, att, AttachmentKind.THESIS_BUNDLE, opposing=is_opposing
                    )
                    changed += 1
                    touched = True
            if touched:
                if is_opposing:
                    self.upsert_opposing_thesis(work)
                else:
                    self.upsert_thesis(work)
        return changed

    def attach_document(
        self,
        thesis_id: str,
        source_path: Path,
        kind: AttachmentKind = AttachmentKind.OTHER,
        label: str | None = None,
        delete_source: bool = False,
    ) -> Attachment:
        """Nahraje soubor do ``~/.bpdpmanager/documents/{thesis_id}/{podsložka}/``.

        Cílový název se generuje podle schématu
        ``{Příjmení}_{typ}_{YYYY-MM-DD}[_vN].{ext}`` (viz ``file_naming``).
        Pokud práce nemá přiřazeného studenta, použije se fallback ``Bez-prijmeni``.
        """
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            raise ValueError(f"Práce {thesis_id} neexistuje.")

        # Příloha (a Jiné) se stejným OBSAHEM už existuje → nepřidávej duplikát
        # (typicky opětovné stažení téhož souboru ze STAG).
        if kind in _DEDUP_KINDS:
            dup = self._find_identical_attachment(
                thesis_id, thesis.attachments, source_path,
                opposing=False, kind=kind,
            )
            if dup is not None:
                dup.is_current = True
                self.upsert_thesis(thesis)
                if delete_source:
                    try:
                        source_path.unlink()
                    except OSError:
                        pass
                return dup

        surname = self._student_surname_for_thesis(thesis)
        subdir = subdir_for(kind)
        target_dir = thesis_documents_dir(thesis_id) / subdir
        target_dir.mkdir(parents=True, exist_ok=True)

        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        target_name = build_target_name(
            surname, kind, source_path, existing_names=existing,
            orig_name=label or source_path.name,
        )
        target_path = target_dir / target_name
        shutil.copy2(source_path, target_path)

        # ``url_or_path`` ukládáme jako relativní cestu vč. podsložky, aby
        # ``document_absolute_path`` fungovala beze změny pro nové i starší záznamy
        # (starší byly v rootu ``documents/{thesis_id}/``, tj. bez podadresáře).
        rel_path = f"{subdir}/{target_name}"

        # Verzování (viz _version_and_supersede): u příloh koexistují různé
        # soubory; verzuje se jen proti souboru stejného názvu.
        label_final = label or target_name
        next_version = self._version_and_supersede(
            thesis.attachments, kind, label_final
        )

        attachment = Attachment(
            label=label_final,
            url_or_path=rel_path,
            kind=kind,
            is_file=True,
            version=next_version,
            is_current=True,
        )
        thesis.attachments.append(attachment)

        # Z nahraného posudku (vedoucího/oponenta) zkus vyčíst navrženou známku.
        # Nový soubor posudku je autoritativní → známka role se PŘEPÍŠE (jinak by
        # dřív špatně vyčtená/stará hodnota držela navždy — smazání ani nové
        # stažení ze STAG by ji neobnovilo). Automatický sync_thesis_grades
        # naproti tomu jen doplňuje prázdné (nepřepisuje ruční úpravy).
        if _is_grade_source(target_path.name):
            grade = extract_grade_from_file(target_path)
            if grade:
                if kind == AttachmentKind.SUPERVISOR_REVIEW:
                    thesis.grade_supervisor = grade
                elif kind == AttachmentKind.OPPONENT_REVIEW:
                    thesis.grade_opponent = grade

        self.upsert_thesis(thesis)

        # Pokud uživatel zaškrtl „smazat originál po nahrání", odstraníme
        # zdrojový soubor — kopie už je bezpečně v ``target_path``.
        # Selhání unlink (např. permission denied) tichá — nemáme co řešit.
        if delete_source:
            try:
                if source_path.is_file():
                    source_path.unlink()
            except OSError:
                pass
        return attachment

    def remove_document(self, thesis_id: str, index: int, delete_file: bool = False) -> None:
        thesis = self.get_thesis(thesis_id)
        if thesis is None or not (0 <= index < len(thesis.attachments)):
            return
        attachment = thesis.attachments[index]
        if delete_file and attachment.is_file:
            target = thesis_documents_dir(thesis_id) / attachment.url_or_path
            if target.exists():
                target.unlink()
        del thesis.attachments[index]
        self.upsert_thesis(thesis)

    def document_absolute_path(self, thesis_id: str, attachment: Attachment) -> Path | None:
        if not attachment.is_file:
            return None
        return thesis_documents_dir(thesis_id) / attachment.url_or_path

    def _student_surname_for_thesis(self, thesis: Thesis) -> str | None:
        """Pomocná funkce — dohledá příjmení studenta navázaného na práci.

        Vrací ``None``, pokud práce nemá studenta nebo student neexistuje
        (``build_target_name`` si pak zařídí fallback).
        """
        if not thesis.student_id:
            return None
        student = self.get_student(thesis.student_id)
        return student.last_name if student else None

    # --- plagiátorství ------------------------------------------------------

    def set_plagiarism_pdf(self, thesis_id: str, source_path: Path) -> str:
        """Zkopíruje PDF s výsledkem plagiátorství do ``plagiat/`` podsložky.

        Uložené ``plagiarism_pdf_filename`` je relativní cesta vč. podadresáře
        (``plagiat/Příjmení_protokol-plagiat_YYYY-MM-DD.pdf``), takže
        ``plagiarism_pdf_path`` funguje stejně pro staré (flat) i nové záznamy.
        """
        from .file_naming import PLAGIARISM_SUBDIR

        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            raise ValueError(f"Práce {thesis_id} neexistuje.")

        surname = self._student_surname_for_thesis(thesis)
        target_dir = thesis_documents_dir(thesis_id) / PLAGIARISM_SUBDIR
        target_dir.mkdir(parents=True, exist_ok=True)

        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        target_name = build_plagiarism_name(surname, source_path, existing_names=existing)
        target_path = target_dir / target_name
        shutil.copy2(source_path, target_path)

        rel_path = f"{PLAGIARISM_SUBDIR}/{target_name}"
        thesis.plagiarism_pdf_filename = rel_path
        self.upsert_thesis(thesis)
        return rel_path

    def remove_plagiarism_pdf(self, thesis_id: str, delete_file: bool = False) -> None:
        thesis = self.get_thesis(thesis_id)
        if thesis is None or not thesis.plagiarism_pdf_filename:
            return
        if delete_file:
            target = thesis_documents_dir(thesis_id) / thesis.plagiarism_pdf_filename
            if target.exists():
                target.unlink()
        thesis.plagiarism_pdf_filename = None
        self.upsert_thesis(thesis)

    def plagiarism_pdf_path(self, thesis_id: str) -> Path | None:
        thesis = self.get_thesis(thesis_id)
        if thesis is None or not thesis.plagiarism_pdf_filename:
            return None
        return thesis_documents_dir(thesis_id) / thesis.plagiarism_pdf_filename

    # --- harmonogram napříč roky --------------------------------------------

    # --- oponentské posudky -------------------------------------------------

    def list_opposing_theses(self) -> list[OpposingThesis]:
        return list(self._db.opposing_theses)

    def get_opposing_thesis(self, op_id: str) -> OpposingThesis | None:
        return next(
            (t for t in self._db.opposing_theses if t.id == op_id), None
        )

    def upsert_opposing_thesis(self, op: OpposingThesis) -> OpposingThesis:
        op.touch()
        existing = self.get_opposing_thesis(op.id)
        if existing:
            idx = self._db.opposing_theses.index(existing)
            self._db.opposing_theses[idx] = op
        else:
            self._db.opposing_theses.append(op)
        self.save()
        return op

    def delete_opposing_thesis(self, op_id: str) -> None:
        self._db.opposing_theses = [
            t for t in self._db.opposing_theses if t.id != op_id
        ]
        self.save()

    def opposing_attach_document(
        self,
        op_id: str,
        source_path: Path,
        kind: AttachmentKind = AttachmentKind.OTHER,
        label: str | None = None,
        delete_source: bool = False,
    ) -> Attachment:
        """Nahraje soubor k oponentskému posudku.

        Soubory leží v ``documents/opposing-{id}/{podsložka}/`` se stejným
        schématem názvu jako u vedených prací — viz ``attach_document``.
        """
        op = self.get_opposing_thesis(op_id)
        if op is None:
            raise ValueError(f"Oponentský posudek {op_id} neexistuje.")

        if kind in _DEDUP_KINDS:
            dup = self._find_identical_attachment(
                op_id, op.attachments, source_path, opposing=True, kind=kind,
            )
            if dup is not None:
                dup.is_current = True
                self.upsert_opposing_thesis(op)
                if delete_source:
                    try:
                        source_path.unlink()
                    except OSError:
                        pass
                return dup

        surname = op.student_last_name or None
        subdir = subdir_for(kind)
        target_dir = thesis_documents_dir(f"opposing-{op_id}") / subdir
        target_dir.mkdir(parents=True, exist_ok=True)

        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        target_name = build_target_name(
            surname, kind, source_path, existing_names=existing,
            orig_name=label or source_path.name,
        )
        target_path = target_dir / target_name
        shutil.copy2(source_path, target_path)

        rel_path = f"{subdir}/{target_name}"

        # Verzování (stejně jako u ``attach_document`` — přílohy koexistují).
        label_final = label or target_name
        next_version = self._version_and_supersede(
            op.attachments, kind, label_final
        )

        attachment = Attachment(
            label=label_final,
            url_or_path=rel_path,
            kind=kind,
            is_file=True,
            version=next_version,
            is_current=True,
        )
        op.attachments.append(attachment)

        # Z nahraného posudku zkus vyčíst navrženou známku — nový soubor je
        # autoritativní, hodnota role se PŘEPÍŠE (vedoucí: externí PDF/Word;
        # oponent: vlastní posudek stažený ze STAG). Automatický
        # sync_opposing_grades jen doplňuje prázdné (nepřepisuje ruční úpravy).
        if _is_grade_source(target_path.name):
            grade = extract_grade_from_file(target_path)
            if grade:
                if kind == AttachmentKind.SUPERVISOR_REVIEW:
                    op.grade_supervisor = grade
                elif kind == AttachmentKind.OPPONENT_REVIEW:
                    op.grade_opponent = grade

        self.upsert_opposing_thesis(op)

        if delete_source:
            try:
                if source_path.is_file():
                    source_path.unlink()
            except OSError:
                pass
        return attachment

    def opposing_remove_document(
        self, op_id: str, index: int, delete_file: bool = False
    ) -> None:
        op = self.get_opposing_thesis(op_id)
        if op is None or not (0 <= index < len(op.attachments)):
            return
        attachment = op.attachments[index]
        if delete_file and attachment.is_file:
            target = thesis_documents_dir(f"opposing-{op_id}") / attachment.url_or_path
            if target.exists():
                target.unlink()
        del op.attachments[index]
        self.upsert_opposing_thesis(op)

    def opposing_document_absolute_path(
        self, op_id: str, attachment: Attachment
    ) -> Path | None:
        if not attachment.is_file:
            return None
        return thesis_documents_dir(f"opposing-{op_id}") / attachment.url_or_path

    def sync_opposing_grades(self, op_id: str) -> OpposingThesis | None:
        """Doplní chybějící známky u oponentury (i zpětně, pro existující data).

        - *Oponent (moje)* ← navržená známka z aktuálního napsaného posudku,
        - *Vedoucí* ← vyčtená z nahraného PDF posudku vedoucího.

        Plní jen prázdné hodnoty (nepřepisuje ručně zadané). Vrací (případně
        aktualizovaný) ``OpposingThesis``.
        """
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return None
        changed = False

        if not op.grade_opponent:
            rev = next(
                (r for r in op.reviews
                 if r.role == "opponent" and r.is_current and r.criteria),
                None,
            )
            if rev is not None:
                op.grade_opponent = rev.suggested_grade
                changed = True
            else:
                # Fallback: vyčti z nahraného posudku oponenta (PDF/Word, typicky
                # vlastní posudek stažený ze STAG, bez napsaného posudku v appce).
                for a in op.attachments:
                    if (
                        a.kind == AttachmentKind.OPPONENT_REVIEW
                        and a.is_file
                        and _is_grade_source(a.url_or_path)
                    ):
                        path = self.opposing_document_absolute_path(op_id, a)
                        if path is not None and path.exists():
                            grade = extract_grade_from_file(path)
                            if grade:
                                op.grade_opponent = grade
                                changed = True
                                break

        if not op.grade_supervisor:
            for a in op.attachments:
                if (
                    a.kind == AttachmentKind.SUPERVISOR_REVIEW
                    and a.is_file
                    and _is_grade_source(a.url_or_path)
                ):
                    path = self.opposing_document_absolute_path(op_id, a)
                    if path is not None and path.exists():
                        grade = extract_grade_from_file(path)
                        if grade:
                            op.grade_supervisor = grade
                            changed = True
                            break

        if changed:
            self.upsert_opposing_thesis(op)
        return op

    def sync_thesis_grades(self, thesis_id: str) -> Thesis | None:
        """Doplní chybějící známky u vedené práce (i zpětně, pro historická data).

        Pro každou roli (vedoucí / oponent):
        - přednost má **in-app posudek** (``Review.suggested_grade`` z kritérií),
        - jinak se zkusí **vyčíst z nahraného posudku** (PDF/Word) dané role.

        Plní jen prázdné hodnoty (nepřepisuje ručně zadané). Vrací (případně
        aktualizovaný) ``Thesis``.
        """
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return None
        changed = False

        roles = (
            ("supervisor", "grade_supervisor", AttachmentKind.SUPERVISOR_REVIEW),
            ("opponent", "grade_opponent", AttachmentKind.OPPONENT_REVIEW),
        )
        for role, field, kind in roles:
            if getattr(thesis, field):
                continue  # už vyplněno (ručně nebo dříve)
            # 1) in-app posudek dané role
            rev = next(
                (r for r in thesis.reviews
                 if r.role == role and r.is_current and r.criteria),
                None,
            )
            if rev is not None:
                setattr(thesis, field, rev.suggested_grade)
                changed = True
                continue
            # 2) fallback: vyčíst z nahraného posudku (PDF/Word)
            for a in thesis.attachments:
                if (
                    a.kind == kind
                    and a.is_file
                    and a.is_current
                    and _is_grade_source(a.url_or_path)
                ):
                    path = self.document_absolute_path(thesis_id, a)
                    if path is not None and path.exists():
                        grade = extract_grade_from_file(path)
                        if grade:
                            setattr(thesis, field, grade)
                            changed = True
                            break

        if changed:
            self.upsert_thesis(thesis)
        return thesis

    # --- posudky k odeslání (PDF poslední verze) ----------------------------

    @staticmethod
    def _current_pdf_attachment(work, kind: AttachmentKind):
        """Vrátí aktuální (is_current) PDF přílohu daného druhu, nebo None."""
        candidates = [
            a
            for a in work.attachments
            if a.kind == kind
            and a.is_file
            and a.url_or_path.lower().endswith(".pdf")
        ]
        if not candidates:
            return None
        # Preferuj is_current; jinak nejvyšší verzi.
        current = [a for a in candidates if a.is_current]
        pool = current or candidates
        return max(pool, key=lambda a: a.version)

    def current_supervisor_review_pdf(self, thesis: Thesis) -> Path | None:
        """Absolutní cesta k aktuálnímu PDF posudku vedoucího, nebo None."""
        att = self._current_pdf_attachment(thesis, AttachmentKind.SUPERVISOR_REVIEW)
        return self.document_absolute_path(thesis.id, att) if att else None

    def current_opponent_review_pdf(self, op: OpposingThesis) -> Path | None:
        """Absolutní cesta k aktuálnímu PDF oponentského posudku, nebo None."""
        att = self._current_pdf_attachment(op, AttachmentKind.OPPONENT_REVIEW)
        return self.opposing_document_absolute_path(op.id, att) if att else None

    def mark_supervisor_review_sent(
        self, thesis_id: str, when: datetime | None = None
    ) -> None:
        """Označí posudek vedoucího jako odeslaný sekretářce."""
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return
        thesis.supervisor_review_sent_at = when or datetime.now()
        self.upsert_thesis(thesis)

    def mark_opponent_review_sent(
        self, op_id: str, when: datetime | None = None
    ) -> None:
        """Označí oponentský posudek jako odeslaný sekretářce."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return
        op.opponent_review_sent_at = when or datetime.now()
        self.upsert_opposing_thesis(op)

    def set_supervisor_review_sent(self, thesis_id: str, sent: bool) -> None:
        """Ručně přepne příznak odeslání posudku vedoucího (ano/ne)."""
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return
        thesis.supervisor_review_sent_at = datetime.now() if sent else None
        self.upsert_thesis(thesis)

    def set_opponent_review_sent(self, op_id: str, sent: bool) -> None:
        """Ručně přepne příznak odeslání oponentského posudku (ano/ne)."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return
        op.opponent_review_sent_at = datetime.now() if sent else None
        self.upsert_opposing_thesis(op)

    def set_supervisor_review_printed(self, thesis_id: str, printed: bool) -> None:
        """Ručně přepne příznak vytištění posudku vedoucího (ano/ne)."""
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return
        thesis.supervisor_review_printed_at = datetime.now() if printed else None
        self.upsert_thesis(thesis)

    def set_opponent_review_printed(self, op_id: str, printed: bool) -> None:
        """Ručně přepne příznak vytištění oponentského posudku (ano/ne)."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return
        op.opponent_review_printed_at = datetime.now() if printed else None
        self.upsert_opposing_thesis(op)

    def auto_link_retakes(self) -> int:
        """Automaticky propojí řádný + opravný pokus (repetent).

        - **Vedené práce:** páruje práce stejného studenta a typu, kde je jedna
          *Obhájeno* a druhá *Nedokončeno* (typický repetent — řádný neobhájen,
          opravný obhájen). Páruje jen dvojice (přesně 2 záznamy).
        - **Oponentury:** páruje dvojice stejného studenta (osobní číslo) a typu.

        Idempotentní. Vrací počet nově spárovaných dvojic.
        """
        linked = 0
        groups: dict[tuple, list[Thesis]] = {}
        for t in self._db.theses:
            if t.student_id:
                groups.setdefault((t.student_id, t.type), []).append(t)
        for grp in groups.values():
            if len(grp) != 2:
                continue
            statuses = {x.status for x in grp}
            if ThesisStatus.DEFENDED in statuses and ThesisStatus.CANCELLED in statuses:
                a, b = grp
                if a.related_thesis_id != b.id or b.related_thesis_id != a.id:
                    a.related_thesis_id, b.related_thesis_id = b.id, a.id
                    a.touch()
                    b.touch()
                    linked += 1

        ogroups: dict[tuple, list[OpposingThesis]] = {}
        for o in self._db.opposing_theses:
            key = (o.student_university_id or o.student_full_name, o.type)
            if key[0]:
                ogroups.setdefault(key, []).append(o)
        for grp in ogroups.values():
            if len(grp) != 2:
                continue
            a, b = grp
            if a.related_thesis_id != b.id or b.related_thesis_id != a.id:
                a.related_thesis_id, b.related_thesis_id = b.id, a.id
                a.touch()
                b.touch()
                linked += 1

        if linked:
            self.save()
        return linked

    # --- harmonogram napříč roky --------------------------------------------

    def upcoming_dates_all_years(self, from_date: date, days: int = 60) -> list[tuple[str, KeyDate]]:
        """Vrátí důležité nadcházející termíny napříč všemi roky."""
        out: list[tuple[str, KeyDate]] = []
        for info in self._db.academic_years:
            for kd in info.upcoming(from_date, days):
                out.append((info.label, kd))
        return sorted(out, key=lambda x: x[1].sort_key())

    # --- knihovna šablon posudků (review templates) -----------------------

    @staticmethod
    def _templates_dir() -> Path:
        """Cesta k podsložce s šablonami v aktuální data_dir."""
        path = app_data_dir() / "templates"
        path.mkdir(parents=True, exist_ok=True)
        return path

    def list_review_templates(
        self,
        *,
        type_filter: ThesisType | None = None,
        role_filter: str | None = None,
        language_filter: str | None = None,
        obor_filter: str | None = None,
    ) -> list[ReviewTemplate]:
        """Vrátí šablony, volitelně filtrované podle metadat.

        Filtry jsou inkluzivní — None znamená „nezohledňovat".
        """
        items = list(self._db.review_templates)
        if type_filter is not None:
            items = [t for t in items if t.type == type_filter]
        if role_filter is not None:
            items = [t for t in items if t.role == role_filter]
        if language_filter is not None:
            items = [t for t in items if t.language == language_filter]
        if obor_filter is not None and obor_filter.strip():
            # Match: prázdný obor v šabloně = univerzální, jinak shoda
            items = [t for t in items if not t.obor or t.obor == obor_filter]
        return sorted(
            items,
            key=lambda t: (t.type.value, t.role, t.language, t.obor, t.academic_year, t.name),
        )

    def get_review_template(self, template_id: str) -> ReviewTemplate | None:
        return next(
            (t for t in self._db.review_templates if t.id == template_id), None
        )

    def review_template_file_path(self, template: ReviewTemplate) -> Path | None:
        """Absolutní cesta k XLSX souboru šablony."""
        if not template.filename:
            return None
        return self._templates_dir() / template.filename

    def register_review_template(
        self,
        name: str,
        type: ThesisType,
        role: str,
        language: str,
        obor: str,
        academic_year: str,
        source_path: Path,
        note: str = "",
    ) -> ReviewTemplate:
        """Zaregistruje XLSX soubor jako šablonu posudku.

        Source soubor se zkopíruje do ``profile_dir/templates/`` pod novým,
        FS-bezpečným názvem, který zahrnuje ID šablony pro odolnost vůči
        duplicitám. Originál se nemodifikuje.
        """
        source_path = Path(source_path)
        if not source_path.is_file():
            raise ValueError(f"Zdrojový soubor neexistuje: {source_path}")
        if source_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Šablona musí být .xlsx (got {source_path.suffix})")

        # Vytvoř ReviewTemplate s předem alokovaným ID, aby cílový název mohl
        # obsahovat krátký prefix ID (robustnější než plný UUID v názvu).
        tmpl = ReviewTemplate(
            name=name.strip() or source_path.stem,
            type=type,
            role=role,
            language=language,
            obor=obor.strip(),
            academic_year=academic_year.strip(),
            note=note.strip(),
        )

        # FS-safe filename: {short_id}_{sanitized name}.xlsx
        short_id = tmpl.id.split("-")[0]
        safe_name = sanitize_for_fs(tmpl.name)[:80] or "template"
        target_name = f"{short_id}_{safe_name}.xlsx"
        target_path = self._templates_dir() / target_name
        # Pokud koliduje (extrémně nepravděpodobné), připoj counter
        n = 2
        while target_path.exists():
            target_name = f"{short_id}_{safe_name}_v{n}.xlsx"
            target_path = self._templates_dir() / target_name
            n += 1
        shutil.copy2(source_path, target_path)

        tmpl.filename = target_name
        self._db.review_templates.append(tmpl)
        self.save()
        return tmpl

    def update_review_template(self, template: ReviewTemplate) -> ReviewTemplate:
        """Re-save metadata existující šablony (nemění filename)."""
        existing = self.get_review_template(template.id)
        if existing is None:
            self._db.review_templates.append(template)
        else:
            idx = self._db.review_templates.index(existing)
            self._db.review_templates[idx] = template
        self.save()
        return template

    def delete_review_template(
        self, template_id: str, *, delete_file: bool = True
    ) -> None:
        """Odstraní šablonu z knihovny + (volitelně) její XLSX soubor."""
        tmpl = self.get_review_template(template_id)
        if tmpl is None:
            return
        if delete_file:
            fp = self.review_template_file_path(tmpl)
            if fp and fp.is_file():
                try:
                    fp.unlink()
                except OSError:
                    pass
        self._db.review_templates = [
            t for t in self._db.review_templates if t.id != template_id
        ]
        self.save()

    def dedupe_review_templates(self, *, dry_run: bool = True) -> dict:
        """Sjednotí form-varianty šablon (prezenční ``-P`` / kombinovaná ``-K``).

        Šablony posudků jsou form-neutrální (forma nehraje roli, jen obor).
        Metoda proto:

        - **sloučí** šablony, které mají shodnou kombinaci typ/role/jazyk/obor
          a **shodný form-neutrální název** (tj. liší se jen značkou ``-P``/
          ``-K`` v názvu) — z každé skupiny ponechá jednu, ostatní odebere;
        - **přejmenuje** ponechané/samostatné šablony na **form-neutrální
          název** (zmizí ``-P``/``-K`` i redundantní ``-EN`` v kódu).

        Vrací ``{"merged": [(odebraná, ponechaná), …],
        "renamed": [(šablona, starý_název, nový_název), …]}``. Při
        ``dry_run=False`` změny provede (odebrané smaže vč. XLSX souboru).
        """
        def _keep_rank(t: ReviewTemplate) -> tuple:
            # Nižší = lepší kandidát na ponechání.
            return (0 if t.criteria else 1, 0 if t.academic_year else 1, t.name.lower())

        groups: dict[tuple, list[ReviewTemplate]] = {}
        for t in self._db.review_templates:
            neutral = form_neutral_name(t.name)
            key = (t.type.value, t.role, t.language, t.obor.strip(), neutral)
            groups.setdefault(key, []).append(t)

        merged: list[tuple[ReviewTemplate, ReviewTemplate]] = []
        renamed: list[tuple[ReviewTemplate, str, str]] = []
        for (_ty, _ro, _la, _ob, neutral), members in groups.items():
            members_sorted = sorted(members, key=_keep_rank)
            keep = members_sorted[0]
            for removed in members_sorted[1:]:
                merged.append((removed, keep))
            if keep.name != neutral:
                renamed.append((keep, keep.name, neutral))

        if not dry_run:
            with self.batch():
                for removed, _keep in merged:
                    self.delete_review_template(removed.id, delete_file=True)
                for keep, _old, new in renamed:
                    fresh = self.get_review_template(keep.id)
                    if fresh is not None:
                        fresh.name = new
                        self.update_review_template(fresh)

        return {"merged": merged, "renamed": renamed}

    # ── vyhledávání prací ─────────────────────────────────────────────────

    @staticmethod
    def _search_fold(s: str) -> str:
        """Pro hledání: bez diakritiky a malými písmeny (``Goláň`` → ``golan``)."""
        nfd = unicodedata.normalize("NFD", s or "")
        return "".join(c for c in nfd if not unicodedata.combining(c)).casefold()

    def search_works(self, query: str) -> list[dict]:
        """Najde vedené práce i oponentury podle jména studenta / názvu / ID / oboru.

        Hledá se **substring, bez ohledu na velikost písmen i diakritiku**
        (``gol`` najde ``Goláň``) ve jméně studenta, názvu práce, univerzitním ID
        (Axxxxx) a oboru. Vrací list dictů ``{kind, id, student, title, status,
        type, uid, obor}``.
        """
        q = self._search_fold((query or "").strip())
        if not q:
            return []
        return [
            h for h in self.search_index()
            if q in self._search_fold(
                f"{h['student']}\n{h['title']}\n{h['uid']}\n{h['obor']}"
            )
        ]

    def search_index(self) -> list[dict]:
        """Vrátí **všechny** práce i oponentury jako hit-dicty pro našeptávač.

        Stejný tvar jako ``search_works`` + navíc ``type`` (``BP`` / ``DP``).
        Slouží jako model pro real-time našeptávání v horním vyhledávání.
        """
        out: list[dict] = []
        for t in self._db.theses:
            student = self.get_student(t.student_id) if t.student_id else None
            name = student.full_name if student else ""
            uid = (student.university_id or "") if student else ""
            obor = (student.obor or "") if student else ""
            out.append({
                "kind": "thesis", "id": t.id, "student": name or "—",
                "title": t.display_title, "status": t.status,
                "type": t.type.value, "uid": uid, "obor": obor,
            })
        for o in self._db.opposing_theses:
            out.append({
                "kind": "opposing", "id": o.id, "student": o.student_full_name or "—",
                "title": o.display_title, "status": None,
                "type": o.type.value, "uid": o.student_university_id or "",
                "obor": o.student_obor or "",
            })
        return out

    # ── výchozí (default) data — obory + šablony ─────────────────────────

    def default_obory_seed_status(self) -> tuple[int, int]:
        """Vrátí ``(chybějící, konfliktní)`` pro výchozí obory.

        - *chybějící* = výchozí obor, který v DB zatím není (podle ``name``),
        - *konfliktní* = existuje pod stejným ``name``, ale s jiným STAG kódem.
        """
        missing = conflicts = 0
        for d in default_obory():
            existing = self.get_obor(d.name)
            if existing is None:
                missing += 1
            elif (existing.stag_code or "") != (d.stag_code or ""):
                conflicts += 1
        return missing, conflicts

    def seed_default_obory(self, *, overwrite_conflicts: bool = False) -> dict[str, int]:
        """Doplní chybějící výchozí obory; volitelně přepíše STAG u konfliktních.

        Existující obory (vč. kontaktů na sekretářku) zůstávají; přepisuje se
        jen ``stag_code`` a jen když ``overwrite_conflicts=True``. Vrací počty
        ``{added, updated, skipped}``.
        """
        added = updated = skipped = 0
        with self.batch():
            for d in default_obory():
                existing = self.get_obor(d.name)
                if existing is None:
                    self._db.obory.append(d.model_copy())
                    added += 1
                elif (existing.stag_code or "") != (d.stag_code or ""):
                    if overwrite_conflicts:
                        existing.stag_code = d.stag_code
                        updated += 1
                    else:
                        skipped += 1
        return {"added": added, "updated": updated, "skipped": skipped}

    def default_template_specs(self) -> list[DefaultTemplateSpec]:
        """Specifikace všech dodávaných šablon (z balíčku)."""
        return list_default_template_specs()

    def default_templates_seed_status(self) -> tuple[int, int]:
        """Vrátí ``(chybějící, existující)`` pro výchozí šablony (podle názvu)."""
        existing_names = {t.name for t in self._db.review_templates}
        missing = present = 0
        for spec in list_default_template_specs():
            if spec.name in existing_names:
                present += 1
            else:
                missing += 1
        return missing, present

    def seed_default_templates(self, *, overwrite_existing: bool = False) -> dict[str, int]:
        """Doplní chybějící výchozí šablony (zkopíruje XLSX + nascanuje schema).

        Při ``overwrite_existing`` se šablona se stejným názvem nejdřív smaže
        a založí znovu z aktuálního zdroje. Vrací ``{added, replaced, skipped}``.
        """
        added = replaced = skipped = 0
        with self.batch():
            for spec in list_default_template_specs():
                existing = next(
                    (t for t in self._db.review_templates if t.name == spec.name),
                    None,
                )
                if existing is not None and not overwrite_existing:
                    skipped += 1
                    continue
                if existing is not None:
                    self.delete_review_template(existing.id, delete_file=True)
                    replaced += 1
                # Akademický rok (a schema) vyčti ze šablony, ať se v manažeru
                # nepropíše prázdný rok.
                academic_year = ""
                try:
                    academic_year = (
                        extract_template_metadata(spec.source_path).get("academic_year")
                        or ""
                    ).strip()
                except Exception:  # noqa: BLE001
                    academic_year = ""
                tmpl = self.register_review_template(
                    name=spec.name,
                    type=spec.type,
                    role=spec.role,
                    language=spec.language,
                    obor=spec.obor,
                    academic_year=academic_year,
                    source_path=spec.source_path,
                )
                try:
                    self.ensure_template_schema(tmpl)
                except Exception:  # noqa: BLE001 — schema necháme dořešit při generování
                    pass
                if existing is None:
                    added += 1
        return {"added": added, "replaced": replaced, "skipped": skipped}

    def reset_obory_to_defaults(self) -> int:
        """Smaže celý číselník oborů a nahradí ho výchozími (vč. STAG kódů).

        Studentům se nemění uložený obor (je to jen string), jen se vyčistí a
        znovu naplní číselník. Vrací počet výchozích oborů.
        """
        self._db.obory = [d.model_copy() for d in default_obory()]
        self.save()
        return len(self._db.obory)

    def reset_templates_to_defaults(self) -> dict[str, int]:
        """Smaže všechny šablony (vč. souborů) a založí znovu výchozí sadu."""
        with self.batch():
            for tmpl in list(self._db.review_templates):
                self.delete_review_template(tmpl.id, delete_file=True)
            res = self.seed_default_templates(overwrite_existing=False)
        return res

    def maybe_seed_defaults(self) -> None:
        """Doseeduje defaulty do právě vytvořeného (nového) profilu.

        Spouští se jen pokud podkladový repozitář právě vytvořil novou DB
        (``created_fresh``) — pro existující profily i v testech je to no-op.
        Obory už naseedovala storage vrstva; tady jen doplníme šablony (a pro
        jistotu i chybějící obory).
        """
        if not getattr(self._repo, "created_fresh", False):
            return
        self._repo.created_fresh = False
        try:
            self.seed_default_obory(overwrite_conflicts=False)
            self.seed_default_templates(overwrite_existing=False)
        except Exception:  # noqa: BLE001 — seed nesmí shodit start aplikace
            pass

    def generate_review_from_template(
        self,
        template_id: str,
        thesis_id: str,
        *,
        attach_as_kind: AttachmentKind | None = None,
    ) -> tuple[Path, Attachment]:
        """Vyplní šablonu daty z práce a připojí ji jako přílohu.

        Vrací ``(absolute_xlsx_path, Attachment)`` — UI může soubor rovnou
        otevřít v Excelu/Numbers.

        ``attach_as_kind`` default odvozen z ``template.role``:
          - supervisor → AttachmentKind.SUPERVISOR_REVIEW
          - opponent → AttachmentKind.OPPONENT_REVIEW
        """
        tmpl = self.get_review_template(template_id)
        if tmpl is None:
            raise ValueError(f"Šablona {template_id} neexistuje.")
        tmpl_path = self.review_template_file_path(tmpl)
        if tmpl_path is None or not tmpl_path.is_file():
            raise ValueError(
                f"Soubor šablony chybí na disku: {tmpl_path}. "
                "Smaž šablonu z knihovny a přidej znovu."
            )

        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            raise ValueError(f"Práce {thesis_id} neexistuje.")

        # Sestav fields
        student = (
            self.get_student(thesis.student_id) if thesis.student_id else None
        )
        opponent = (
            self.get_opponent(thesis.opponent_id) if thesis.opponent_id else None
        )
        user_name = self._guess_user_name()

        if attach_as_kind is None:
            attach_as_kind = (
                AttachmentKind.SUPERVISOR_REVIEW
                if tmpl.role == "supervisor"
                else AttachmentKind.OPPONENT_REVIEW
            )

        student_label = student.full_name if student else ""

        # Při vyplňování: pokud uživatel je oponent, „opponent" je on (user_name),
        # „supervisor" je oponent.name z práce (= cizí vedoucí). Ale tady jsme
        # u vedených prací (Thesis), takže supervisor=user_name, opponent=opp.name.
        fields = {
            "student": student_label,
            "supervisor": user_name if tmpl.role == "supervisor" else "",
            "opponent": user_name if tmpl.role == "opponent" else (
                opponent.name if opponent else ""
            ),
            "title_cs": thesis.title_cs,
            "title_en": thesis.title_en,
            "academic_year": thesis.academic_year,
        }

        # Cíl: do documents/{thesis_id}/posudky/ pod jednotným názvem
        surname = self._student_surname_for_thesis(thesis)
        subdir = subdir_for(attach_as_kind)
        target_dir = thesis_documents_dir(thesis_id) / subdir
        target_dir.mkdir(parents=True, exist_ok=True)
        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        target_name = build_target_name(
            surname, attach_as_kind, tmpl_path, existing_names=existing
        )
        target_path = target_dir / target_name

        # Vyplnění + zápis
        fill_template(tmpl_path, target_path, fields)

        # Zaregistruj jako Attachment přes attach_document-like flow,
        # ale soubor už jsme zapsali přímo do cílové cesty — proto použijeme
        # přímo append + auto-versioning ručně.
        same_kind = [a for a in thesis.attachments if a.kind == attach_as_kind]
        next_version = max((a.version for a in same_kind), default=0) + 1
        for a in same_kind:
            a.is_current = False

        rel_path = f"{subdir}/{target_name}"
        attachment = Attachment(
            label=target_name,
            url_or_path=rel_path,
            kind=attach_as_kind,
            is_file=True,
            version=next_version,
            is_current=True,
        )
        thesis.attachments.append(attachment)
        self.upsert_thesis(thesis)
        return target_path, attachment

    @staticmethod
    def _active_profile_dict() -> dict:
        """Přečte (read-only) dict aktivního profilu z ProfileRegistry.

        Services nezávisí na UI/ProfileManageru, takže registry čteme přímo
        ze souboru. Vrací prázdný dict, pokud nic není.
        """
        try:
            import json

            from ..config import profiles_registry_path

            p = profiles_registry_path()
            if p.is_file():
                data = json.loads(p.read_text(encoding="utf-8"))
                last_id = data.get("last_opened")
                for prof in data.get("profiles", []) or []:
                    if isinstance(prof, dict) and prof.get("id") == last_id:
                        return prof
        except Exception:  # noqa: BLE001
            pass
        return {}

    def _guess_user_name(self) -> str:
        """Jméno uživatele aktivního profilu (bez titulů). Prázdné, když není."""
        return (self._active_profile_dict().get("user_name") or "").strip()

    def _guess_user_titles(self) -> tuple[str, str]:
        """Tituly před/za jménem uživatele aktivního profilu."""
        prof = self._active_profile_dict()
        return (
            (prof.get("user_title_before") or "").strip(),
            (prof.get("user_title_after") or "").strip(),
        )

    def review_author_name(self) -> str:
        """Jméno autora posudku (uživatel profilu) vč. titulů před/za.

        Toto se propisuje do buňky „Vedoucí"/„Oponent" v XLSX posudku.
        """
        from ..models.naming import compose_titled_name

        before, after = self._guess_user_titles()
        return compose_titled_name(before, self._guess_user_name(), after)

    def _guess_review_place(self) -> str:
        """Místo pro podpisový blok posudku z aktivního profilu (default „Zlín")."""
        place = (self._active_profile_dict().get("review_place") or "").strip()
        return place or "Zlín"

    @staticmethod
    def build_place_date(place: str) -> str:
        """Sestaví „Místo, D. M. YYYY" s aktuálním datem (české formátování)."""
        from datetime import date as _date

        today = _date.today()
        date_str = f"{today.day}. {today.month}. {today.year}"
        place = (place or "Zlín").strip()
        return f"{place}, {date_str}"

    # --- review (strukturovaný posudek) -----------------------------------

    def ensure_template_schema(self, tmpl: ReviewTemplate) -> ReviewTemplate:
        """Pokud šablona nemá nascanované schema kritérií, doplní ho.

        Persistuje update do db.json.
        """
        if tmpl.criteria:
            return tmpl  # už máme
        fp = self.review_template_file_path(tmpl)
        if fp is None or not fp.is_file():
            return tmpl
        try:
            schema = extract_template_schema(fp)
        except Exception:  # noqa: BLE001
            return tmpl
        tmpl.criteria = [TemplateCriterion(**c) for c in schema["criteria"]]
        tmpl.field_cells = schema["field_cells"]
        self.update_review_template(tmpl)
        return tmpl

    def list_reviews(
        self, thesis_id: str, *, opposing: bool = False
    ) -> list[Review]:
        """Vrátí všechny ``Review`` k dané práci (nebo oponentskému posudku)."""
        if opposing:
            op = self.get_opposing_thesis(thesis_id)
            return list(op.reviews) if op else []
        t = self.get_thesis(thesis_id)
        return list(t.reviews) if t else []

    def get_current_review(
        self, thesis_id: str, role: str, *, opposing: bool = False
    ) -> Review | None:
        """Vrátí aktuální (is_current) ``Review`` daného role pro práci."""
        reviews = self.list_reviews(thesis_id, opposing=opposing)
        for r in sorted(reviews, key=lambda x: x.version, reverse=True):
            if r.role == role and r.is_current:
                return r
        return None

    def upsert_review(
        self,
        thesis_id: str,
        review: Review,
        *,
        opposing: bool = False,
    ) -> Review:
        """Vloží/aktualizuje ``Review`` v Thesis nebo OpposingThesis.

        Verzování: pokud Review s daným ``id`` existuje → update in-place.
        Jinak append + setify is_current na aktuální (pro daný role).
        """
        review.touch()
        if opposing:
            op = self.get_opposing_thesis(thesis_id)
            if op is None:
                raise ValueError(f"Oponentský posudek {thesis_id} neexistuje.")
            container_list = op.reviews
        else:
            t = self.get_thesis(thesis_id)
            if t is None:
                raise ValueError(f"Práce {thesis_id} neexistuje.")
            container_list = t.reviews

        # Update or insert
        existing = next((r for r in container_list if r.id == review.id), None)
        if existing is not None:
            idx = container_list.index(existing)
            container_list[idx] = review
        else:
            container_list.append(review)
            # Auto-versioning: pokud je is_current=True, ostatní téhož role
            # se přepnou na False (nová verze nahrazuje předchozí current).
            if review.is_current:
                for r in container_list:
                    if r.id != review.id and r.role == review.role:
                        r.is_current = False

        if opposing:
            # Známku v OpposingThesis automaticky doplň z napsaného posudku
            # (oponentova z opponent role, vedoucího z případné supervisor role).
            if review.is_current and review.criteria:
                if review.role == "opponent":
                    op.grade_opponent = review.suggested_grade
                elif review.role == "supervisor":
                    op.grade_supervisor = review.suggested_grade
            self.upsert_opposing_thesis(op)
        else:
            self.upsert_thesis(t)
        return review

    def delete_review(
        self, thesis_id: str, review_id: str, *, opposing: bool = False
    ) -> None:
        """Odstraní ``Review`` z DB (soubory XLSX/PDF zůstávají jako Attachment)."""
        if opposing:
            op = self.get_opposing_thesis(thesis_id)
            if op is None:
                return
            op.reviews = [r for r in op.reviews if r.id != review_id]
            self.upsert_opposing_thesis(op)
        else:
            t = self.get_thesis(thesis_id)
            if t is None:
                return
            t.reviews = [r for r in t.reviews if r.id != review_id]
            self.upsert_thesis(t)

    def _relink_review_template(
        self,
        review: Review,
        *,
        expected_type: ThesisType | None = None,
        obor_hint: str = "",
    ) -> ReviewTemplate | None:
        """Dohledá aktuální šablonu pro posudek, jehož ``template_id`` už nesedí.

        ID se mění např. po „Smazat vše a nahradit" v knihovně šablon. Pořadí:

        1. přesně podle uloženého názvu (``template_name``),
        2. jinak podle **role (+ typ práce + jazyk + obor)** — vybere se nejlepší
           současná šablona; pokud existuje aspoň jedna vhodná, nikdy neselže.

        ``expected_type`` a ``obor_hint`` (kód disciplíny, např. „SWI") předává
        ``generate_review_files`` z kontextu práce. Posudek se přepojí
        (přepíše se ``template_id`` i ``template_name``).
        """
        templates = self._db.review_templates
        # 1) přesně podle uloženého názvu
        if review.template_name:
            for t in templates:
                if t.name == review.template_name:
                    review.template_id = t.id
                    return t
        # 2) heuristika podle kontextu práce
        cands = [t for t in templates if t.role == review.role]
        if expected_type is not None:
            typed = [t for t in cands if t.type == expected_type]
            cands = typed or cands
        lang = [t for t in cands if t.language == review.language]
        cands = lang or cands
        if not cands:
            return None
        if obor_hint:
            by_obor = [
                t for t in cands if t.obor and t.obor.upper() == obor_hint.upper()
            ]
            if by_obor:
                cands = by_obor
        if review.criteria:
            n = len(review.criteria)
            by_crit = [t for t in cands if t.criteria and len(t.criteria) == n]
            if by_crit:
                cands = by_crit
        chosen = cands[0]
        review.template_id = chosen.id
        review.template_name = chosen.name
        return chosen

    def generate_review_files(
        self,
        thesis_id: str,
        review: Review,
        *,
        opposing: bool = False,
        also_pdf: bool = True,
    ) -> tuple[Path, Path | None]:
        """Vyrenderuje XLSX a (volitelně) PDF z ``Review`` dat + šablony.

        - Otevře šablonu, vyplní základní pole + kritéria + extra pole
        - Uloží jako příloha typu SUPERVISOR_REVIEW / OPPONENT_REVIEW
          (auto-versioning v Attachment + Review)
        - Pokud ``also_pdf=True`` a LibreOffice je k dispozici, vyrobí
          i PDF a uloží jako další přílohu téhož kindu

        Returns ``(xlsx_path, pdf_path_or_None)``.
        """
        tmpl = self.get_review_template(review.template_id) if review.template_id else None
        if tmpl is None:
            # Stalé ID (např. po „Smazat vše a nahradit" / přegenerování
            # defaultů) — přepoj podle názvu, jinak podle kontextu práce.
            exp_type: ThesisType | None = None
            obor_hint = ""
            if opposing:
                _w = self.get_opposing_thesis(thesis_id)
                if _w is not None:
                    exp_type = _w.type
                    obor_hint = discipline_from_app_code(_w.student_obor or "")
            else:
                _w = self.get_thesis(thesis_id)
                if _w is not None:
                    exp_type = _w.type
                    _st = self.get_student(_w.student_id) if _w.student_id else None
                    obor_hint = discipline_from_app_code((_st.obor if _st else "") or "")
            tmpl = self._relink_review_template(
                review, expected_type=exp_type, obor_hint=obor_hint
            )
        if tmpl is None:
            raise ValueError(
                "Šablona posudku nebyla nalezena. Pravděpodobně byla smazána "
                "z knihovny — přidej ji znovu nebo zvol jinou."
            )
        tmpl = self.ensure_template_schema(tmpl)
        tmpl_path = self.review_template_file_path(tmpl)
        if tmpl_path is None or not tmpl_path.is_file():
            raise ValueError(f"Soubor šablony chybí na disku: {tmpl_path}")

        # Cíl
        if opposing:
            container_id = f"opposing-{thesis_id}"
            opp = self.get_opposing_thesis(thesis_id)
            if opp is None:
                raise ValueError(f"Oponentský posudek {thesis_id} neexistuje.")
            surname = opp.student_last_name or None
        else:
            container_id = thesis_id
            t = self.get_thesis(thesis_id)
            if t is None:
                raise ValueError(f"Práce {thesis_id} neexistuje.")
            surname = self._student_surname_for_thesis(t)

        kind = (
            AttachmentKind.SUPERVISOR_REVIEW
            if review.role == "supervisor"
            else AttachmentKind.OPPONENT_REVIEW
        )
        subdir = subdir_for(kind)
        target_dir = thesis_documents_dir(container_id) / subdir
        target_dir.mkdir(parents=True, exist_ok=True)

        # Archivace předchozích posudků téhož typu: starší XLSX se přesunou
        # do ``posudky/archiv/`` (přejmenované s timestampem), stará PDF se
        # smažou (jsou jen odvozeninou). Po tomto kroku zůstává „1 aktuální +
        # archiv" a nový posudek dostane čistý název bez ``_vN``.
        atts = opp.attachments if opposing else t.attachments
        self._archive_previous_review_files(container_id, atts, kind, subdir)

        existing = {p.name for p in target_dir.iterdir() if p.is_file()}
        target_name = build_target_name(
            surname, kind, tmpl_path, existing_names=existing
        )
        target_path = target_dir / target_name

        # === Vyplnění buněk BEZ ztráty zbytku šablony (logo, kresby, styly) ===
        # KRITICKÉ: nepřepisujeme celý sešit přes openpyxl (to zahazuje
        # obrázky v záhlaví / logo), ale zapíšeme jen hodnoty buněk přes
        # xlsx_cell_writer (XML aktivního listu) — výstup je 1:1 se šablonou.

        # 1) Naplánuj základní pole (heuristika dle popisků sloupce A) — readonly
        basic_fields = {
            "student": review.student_name,
            "supervisor": review.user_name if review.role == "supervisor" else "",
            "opponent": review.user_name if review.role == "opponent" else "",
            "title_cs": review.title_cs,
            "title_en": review.title_en,
            "academic_year": review.academic_year,
        }
        values: dict[str, object] = {}
        try:
            for coord, _fkey, value in plan_template_fill(tmpl_path, basic_fields):
                values[coord] = value
        except Exception:  # noqa: BLE001
            pass

        # 2) Kritéria — body + (volitelně) váhy
        for cs in review.criteria:
            if cs.score_cell:
                try:
                    values[cs.score_cell] = float(cs.score)
                except (TypeError, ValueError):
                    pass
            if cs.weight_cell:
                try:
                    values[cs.weight_cell] = float(cs.weight)
                except (TypeError, ValueError):
                    pass

        # 3) Extra pole — place_date vyžaduje znát původní text buňky šablony
        #    (obsahuje „Místo, datum: …  Podpis: …" v jednom textu).
        place_existing = self._template_cell_text(
            tmpl_path, tmpl.field_cells.get("place_date")
        )
        extras = {
            "assignment_fulfilled": review.assignment_fulfilled,
            "plagiarism_verdict": review.plagiarism_verdict,
            "plagiarism_justification": review.plagiarism_justification,
            "overall_comment": review.overall_comment,
            "place_date": review.place_date,
        }
        for key, value in extras.items():
            cell = tmpl.field_cells.get(key)
            if not cell or not value:
                continue
            if key == "place_date":
                # Nahraď jen PRVNÍ tečkovanou linku (za „Místo, datum:")
                # hodnotou a podpisový blok ponech nedotčený.
                if isinstance(place_existing, str) and "..." in place_existing:
                    values[cell] = re.sub(
                        r"\.{3,}", f" {value} ", place_existing, count=1
                    )
                elif isinstance(place_existing, str) and place_existing.strip():
                    values[cell] = f"{place_existing.rstrip()} {value}"
                else:
                    values[cell] = value
            else:
                values[cell] = value

        # 4) Dlouhé volné texty (slovní hodnocení, zdůvodnění): nastav výšku
        #    řádku, a když se text nevejde na stránku, rozděl sloučenou buňku
        #    na víc řádků (jen když je to nutné) — ať se v PDF neusekne.
        row_heights: dict[int, float] = {}
        merge_splits: dict[str, list[str]] = {}
        cell_styles: dict[str, str] = {}
        for key in ("overall_comment", "plagiarism_justification",
                    "assignment_fulfilled"):
            cell = tmpl.field_cells.get(key)
            text = extras.get(key)
            if not cell or not text:
                continue
            vals, splits, heights, styles = self._plan_long_text(
                tmpl_path, cell, str(text)
            )
            values.pop(cell, None)  # nahradíme případně víc buňkami
            values.update(vals)
            merge_splits.update(splits)
            row_heights.update(heights)
            cell_styles.update(styles)

        # 5) Jeden zápis — zachová logo i veškeré formátování šablony 1:1
        set_cells(
            tmpl_path, target_path, values,
            row_heights=row_heights, merge_splits=merge_splits,
            cell_styles=cell_styles,
        )

        # 3) Připoj XLSX jako Attachment (auto-versioning)
        same_kind = []
        if opposing:
            same_kind = [a for a in opp.attachments if a.kind == kind]
        else:
            same_kind = [a for a in t.attachments if a.kind == kind]
        next_version = max((a.version for a in same_kind), default=0) + 1
        for a in same_kind:
            a.is_current = False

        rel_path = f"{subdir}/{target_name}"
        xlsx_att = Attachment(
            label=target_name,
            url_or_path=rel_path,
            kind=kind,
            is_file=True,
            version=next_version,
            is_current=True,
        )
        if opposing:
            opp.attachments.append(xlsx_att)
        else:
            t.attachments.append(xlsx_att)
        review.xlsx_filename = rel_path

        # 4) PDF přes LibreOffice (volitelné)
        pdf_path: Path | None = None
        if also_pdf:
            pdf_path = self._xlsx_to_pdf(target_path)
            if pdf_path is not None:
                # Připoj PDF jako další attachment téhož kindu
                rel_pdf = f"{subdir}/{pdf_path.name}"
                pdf_att = Attachment(
                    label=pdf_path.name,
                    url_or_path=rel_pdf,
                    kind=kind,
                    is_file=True,
                    version=next_version,  # stejná verze jako XLSX
                    # PDF i XLSX nejnovějšího posudku jsou „aktuální" — ať se
                    # PDF ukazuje v seznamu hned, ne až po zobrazení starších.
                    is_current=True,
                )
                if opposing:
                    opp.attachments.append(pdf_att)
                else:
                    t.attachments.append(pdf_att)
                review.pdf_filename = rel_pdf

        # 5) Persist
        if opposing:
            self.upsert_opposing_thesis(opp)
        else:
            self.upsert_thesis(t)

        return target_path, pdf_path

    def _archive_previous_review_files(
        self,
        container_id: str,
        attachments: list[Attachment],
        kind: AttachmentKind,
        subdir: str,
    ) -> None:
        """Uklidí předchozí soubory posudku téhož ``kind`` před generováním nového.

        - starší XLSX (a jiné než PDF) se přesunou do ``{subdir}/archiv/`` a
          přejmenují na ``{stem}_archiv_{YYYY-MM-DD_HHMMSS}.xlsx``; zůstávají
          jako přílohy (``is_current=False``), aby k nim šlo dohledat historii,
        - starší PDF se smažou i s odpovídající přílohou (PDF je odvozenina XLSX).

        Mutuje ``attachments`` in-place (typicky ``thesis.attachments`` /
        ``op.attachments``). Posudky druhé role (např. oponent při generování
        vedoucího) zůstávají nedotčené — filtrujeme striktně na ``kind``.
        """
        base_dir = thesis_documents_dir(container_id)
        archive_dir = base_dir / subdir / "archiv"
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")

        kept: list[Attachment] = []
        for att in attachments:
            # Soubory, které UŽ JSOU v archivu, nech být — jinak by se jim při
            # každém novém posudku znovu přidával „_archiv_" do názvu (kumulace).
            already_archived = Path(att.url_or_path).parent.name == "archiv"
            if att.kind != kind or not att.is_file or already_archived:
                kept.append(att)
                continue
            abs_path = base_dir / att.url_or_path
            suffix = Path(att.url_or_path).suffix.lower()
            if suffix == ".pdf":
                # Stará PDF je jen odvozenina — smaž soubor i záznam.
                try:
                    if abs_path.exists():
                        abs_path.unlink()
                except OSError:
                    pass
                continue  # přílohu nezachováme
            # XLSX (nebo jiný formát posudku) → přesun do archivu.
            att.is_current = False
            if abs_path.exists():
                archive_dir.mkdir(parents=True, exist_ok=True)
                archived_name = self._unique_archive_name(
                    archive_dir, abs_path.stem, ts, abs_path.suffix
                )
                try:
                    shutil.move(str(abs_path), str(archive_dir / archived_name))
                    att.url_or_path = f"{subdir}/archiv/{archived_name}"
                    att.label = archived_name
                except OSError:
                    pass  # ponech původní cestu, ať nezmizí ze seznamu
            kept.append(att)

        attachments[:] = kept

    @staticmethod
    def _unique_archive_name(
        archive_dir: Path, stem: str, ts: str, suffix: str
    ) -> str:
        """Sestaví unikátní název v ``archiv/`` (timestamp + příp. pořadí)."""
        candidate = f"{stem}_archiv_{ts}{suffix}"
        if not (archive_dir / candidate).exists():
            return candidate
        n = 2
        while True:
            candidate = f"{stem}_archiv_{ts}_{n}{suffix}"
            if not (archive_dir / candidate).exists():
                return candidate
            n += 1

    # Vícenásobně zanořené „_archiv_<datum>_<čas>" segmenty (důsledek staršího
    # bugu) — sloučíme je na první výskyt.
    _NESTED_ARCHIVE_RE = re.compile(
        r"(_archiv_\d{4}-\d{2}-\d{2}_\d{6}(?:_\d+)?)"
        r"(?:_archiv_\d{4}-\d{2}-\d{2}_\d{6}(?:_\d+)?)+"
    )

    # --- komise SZZ ---------------------------------------------------------

    def list_committees(self) -> list:
        return list(self._db.committees)

    def get_committee(self, committee_id: str):
        return next((c for c in self._db.committees if c.id == committee_id), None)

    def upsert_committee(self, committee) -> None:
        db = self._db
        for i, c in enumerate(db.committees):
            if c.id == committee.id:
                db.committees[i] = committee
                self.save()
                return
        db.committees.append(committee)
        self.save()

    def delete_committee(self, committee_id: str) -> None:
        db = self._db
        db.committees = [c for c in db.committees if c.id != committee_id]
        self.save()

    @staticmethod
    def _komise_seed_path() -> Path:
        """Cesta ke kurátorovanému JSONu komisí (veřejná data v gitu)."""
        return (
            Path(__file__).resolve().parent.parent / "resources" / "komise_szz.json"
        )

    def load_komise_seed(self) -> dict:
        """Načte komise z kurátorovaného JSONu a sloučí je do databáze.

        Složení komisí (barva, obor, stupeň, členové, data) je *veřejné* a
        verzované v gitu (``resources/komise_szz.json``). Sloty studentů
        (jména/osobní čísla) v JSONu NEJSOU — ty se plní jen lokálně z PDF
        rozpisů. Funkce je idempotentní; volá se při startu aplikace.

        Slučovací klíč: (rok, stupeň, obor, barva). Když starší lokálně
        naimportovaná komise nemá obor (data před 2.5.0), doplní se z JSONu
        (match dle rok+stupeň+barva, právě jeden kandidát bez oboru) — místo
        vytvoření duplikátu. Sloty, source_files ani uživatelské poznámky se
        nepřepisují; složení (členové, data, program, obor) se aktualizuje.
        """
        import json

        from ..models import Committee, CommitteeMember

        path = self._komise_seed_path()
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, ValueError):
            return {"created": 0, "updated": 0}

        db = self._db
        stats = {"created": 0, "updated": 0}

        def find(year: str, level: str, obor: str, color: str):
            # 1) přesná shoda vč. oboru
            for c in db.committees:
                if (c.academic_year == year and c.level == level
                        and c.obor == obor and c.color == color):
                    return c
            # 2) starší komise bez oboru (rok+stupeň+barva) — právě jedna
            cands = [
                c for c in db.committees
                if c.academic_year == year and c.level == level
                and c.color == color and not c.obor
            ]
            return cands[0] if len(cands) == 1 else None

        for entry in data.get("committees", []):
            year = entry.get("academic_year", "")
            level = entry.get("level", "")
            obor = entry.get("obor", "")
            color = entry.get("color", "")
            members = [
                CommitteeMember(role=m.get("role", "Člen"), name=m.get("name", ""))
                for m in entry.get("members", [])
            ]
            c = find(year, level, obor, color)
            if c is None:
                c = Committee(academic_year=year, level=level, color=color)
                db.committees.append(c)
                stats["created"] += 1
            else:
                stats["updated"] += 1
            # Složení je z JSONu autoritativní — přepiš ho; sloty/zdroje nech být.
            c.obor = obor
            c.program_label = entry.get("program_label", "")
            c.members = members
            c.note = entry.get("note", "")
            c.from_seed = True
            for d in entry.get("dates", []):
                if d not in c.dates:
                    c.dates.append(d)

        self.save()
        return stats

    def reset_committees_from_seed(self) -> dict:
        """Smaže všechny komise a načte je znovu z kurátorovaného JSONu.

        Úklid starých (před 2.5.0) lokálně naimportovaných komisí, které
        „nesedí" (chybí obor, duplicity, zmíchané barvy). **Sloty studentů
        z dříve nahraných PDF se ztratí** — rozpisy je potřeba naimportovat
        znovu (napojí se už na správné komise). Vrací statistiku seedu.
        """
        self._db.committees = []
        return self.load_komise_seed()

    def komise_store_pdf(self, source: Path, academic_year: str, *,
                         name: str | None = None, kind: str = "rozpisy") -> str:
        """Zkopíruje PDF do ``komise/<rok>/<kind>/`` (strukturované uložení).

        ``kind`` = ``"rozpisy"`` (rozpisy studentů) / ``"slozeni"`` (složení
        komisí). ``name`` = cílový (přejmenovaný) název souboru; když chybí,
        použije se původní. Vrací relativní cestu (vůči složce komise).
        Stejnojmenný soubor se přepíše (novější verze nahradí starší).
        """
        import shutil

        from ..config import komise_dir

        safe_year = (academic_year or "ostatni").replace("/", "-")
        sub = kind if kind in ("rozpisy", "slozeni") else "rozpisy"
        target_dir = komise_dir() / safe_year / sub
        target_dir.mkdir(parents=True, exist_ok=True)
        fname = self._safe_pdf_name(name or Path(source).name)
        target = target_dir / fname
        shutil.copy2(source, target)
        return f"{safe_year}/{sub}/{target.name}"

    @staticmethod
    def _safe_pdf_name(name: str) -> str:
        """Bezpečný název PDF (bez cest/divných znaků), s příponou .pdf."""
        stem = re.sub(r"[^0-9A-Za-zÀ-ž._+-]+", "-", Path(name).stem).strip("-._")
        return (stem or "soubor") + ".pdf"

    def komise_pdf_path(self, rel_path: str) -> Path:
        from ..config import komise_dir

        return komise_dir() / rel_path

    @staticmethod
    def _komise_seed_pdf_dir() -> Path:
        """Složka s PDF složení komisí dodanými v gitu (veřejná data)."""
        return Path(__file__).resolve().parent.parent / "resources" / "komise_pdfs"

    def komise_pdf_inventory(self) -> dict:
        """Seznam PDF komisí po akademických rocích → {rok: {slozeni, rozpisy}}.

        Slučuje **PDF složení dodaná v gitu** (``resources/komise_pdfs/<rok>/``)
        s **lokálně nahranými** (``komise/<rok>/slozeni|rozpisy/``; starší PDF
        přímo v ``komise/<rok>/`` se berou jako rozpisy). Hodnoty jsou seznamy
        absolutních cest, deduplikované podle názvu souboru.
        """
        from ..config import komise_dir

        inv: dict[str, dict[str, list[Path]]] = {}

        def slot(year: str) -> dict[str, list[Path]]:
            return inv.setdefault(
                year, {"slozeni": [], "rozpisy": [], "nezarazene": []}
            )

        def add(year: str, key: str, pdf: Path) -> None:
            bucket = slot(year)[key]
            if pdf.name not in {p.name for p in bucket}:
                bucket.append(pdf)

        # 1) Složení dodané v gitu.
        seed = self._komise_seed_pdf_dir()
        if seed.exists():
            for ydir in sorted(seed.iterdir()):
                if ydir.is_dir():
                    year = ydir.name.replace("-", "/")
                    for pdf in sorted(ydir.glob("*.pdf")):
                        add(year, "slozeni", pdf)
        # 2) Lokálně nahraná PDF.
        kd = komise_dir()
        if kd.exists():
            for ydir in sorted(kd.iterdir()):
                if not ydir.is_dir():
                    continue
                year = ydir.name.replace("-", "/")
                for sub in ("slozeni", "rozpisy"):
                    subdir = ydir / sub
                    if subdir.exists():
                        for pdf in sorted(subdir.glob("*.pdf")):
                            add(year, sub, pdf)
                # Starší PDF přímo v komise/<rok>/ (před 2.5.5) — nezařazené
                # (z první verze importu, často se špatným názvem). Uživatel je
                # smaže z panelu nebo nahradí novým importem.
                for pdf in sorted(ydir.glob("*.pdf")):
                    add(year, "nezarazene", pdf)
        return inv

    def komise_delete_pdf(self, abs_path) -> bool:
        """Smaže PDF soubor z disku — jen pokud leží ve složce ``komise/``.

        PDF dodaná v gitu (``resources/komise_pdfs/``) se mazat nedají (vrátí
        ``False``). Vrací ``True`` při úspěšném smazání.
        """
        from ..config import komise_dir

        p = Path(abs_path).resolve()
        try:
            p.relative_to(komise_dir().resolve())
        except ValueError:
            return False   # není v komise/ (např. dodané v gitu) → nemazat
        try:
            p.unlink(missing_ok=True)
            return True
        except OSError:
            return False

    def apply_komise_import(self, committees, schedules,
                            source_rel_paths: list[str]) -> dict:
        """Zapíše naparsované komise a rozpisy do databáze (s merge).

        - Komise se slučují podle (rok, stupeň, OBOR, barva): členové
          sjednocení podle jména, data sjednocená, delší program label vyhrává.
          Obor je v klíči nutný — sama barva nestačí (Mgr fialová je NKYB i NUI).
        - Rozpis se připojí ke komisi stejného klíče (typicky seed komise
          z JSONu); sloty stejného (datum, čas, os. číslo) se nepřidají podruhé.
        """
        from ..models import Committee, CommitteeMember, DefenseSlot

        db = self._db
        stats = {"created": 0, "updated": 0, "slots": 0}

        def find(year: str, level: str, obor: str, color: str):
            def lvl_ok(c) -> bool:
                return c.level == level or not c.level or not level
            # 1) Přesná shoda včetně oboru (typicky seed komise z JSONu).
            if obor:
                for c in db.committees:
                    if (c.academic_year == year and c.color == color
                            and lvl_ok(c) and c.obor == obor):
                        return c
            # 2) Starší komise bez oboru (data před 2.5.0) — právě jedna.
            no_obor = [
                c for c in db.committees
                if c.academic_year == year and c.color == color
                and lvl_ok(c) and not c.obor
            ]
            if len(no_obor) == 1:
                return no_obor[0]
            # 3) Obor se nepodařilo určit → shoda dle rok+stupeň+barva, je-li jediná.
            if not obor:
                any_match = [
                    c for c in db.committees
                    if c.academic_year == year and c.color == color and lvl_ok(c)
                ]
                if len(any_match) == 1:
                    return any_match[0]
            return None

        def touch_sources(c) -> None:
            for rp in source_rel_paths:
                if rp not in c.source_files:
                    c.source_files.append(rp)

        for pc in committees:
            obor = getattr(pc, "obor", "")
            c = find(pc.academic_year, pc.level, obor, pc.color)
            if c is None:
                c = Committee(academic_year=pc.academic_year, color=pc.color,
                              level=pc.level, obor=obor)
                db.committees.append(c)
                stats["created"] += 1
            else:
                stats["updated"] += 1
            if not c.obor and obor:
                c.obor = obor
            for d in pc.dates:
                if d not in c.dates:
                    c.dates.append(d)
            # Seed komise (z JSONu) má složení autoritativní — z PDF jen sloty
            # a zdroje; program/členy/úroveň nepřepisuj.
            if not c.from_seed:
                if len(pc.program_label) > len(c.program_label):
                    c.program_label = pc.program_label
                if not c.level and pc.level:
                    c.level = pc.level
                existing = {m.name for m in c.members}
                for role, name in pc.members:
                    if name not in existing:
                        c.members.append(CommitteeMember(role=role, name=name))
                        existing.add(name)
            touch_sources(c)

        for ps in schedules:
            obor = getattr(ps, "obor", "")
            c = find(ps.academic_year, ps.level, obor, ps.color)
            if c is None:
                c = Committee(academic_year=ps.academic_year, color=ps.color,
                              level=ps.level, obor=obor,
                              program_label=ps.program_label)
                db.committees.append(c)
                stats["created"] += 1
            if not c.obor and obor:
                c.obor = obor
            # Seed komise má program label z JSONu — z rozpisu ho nepřepisuj.
            if not c.from_seed and len(ps.program_label) > len(c.program_label):
                c.program_label = ps.program_label
            for d in ps.dates:
                if d not in c.dates:
                    c.dates.append(d)
            seen = {(s.date, s.time, s.personal_number) for s in c.slots}
            for slot_date, slot_time, pnum, name in ps.slots:
                if (slot_date, slot_time, pnum) not in seen:
                    c.slots.append(DefenseSlot(date=slot_date, time=slot_time,
                                               personal_number=pnum,
                                               student_name=name))
                    seen.add((slot_date, slot_time, pnum))
                    stats["slots"] += 1
            touch_sources(c)

        self.save()
        return stats

    def komise_student_roles(self) -> dict[str, str]:
        """Mapa pro zvýraznění v rozpisech: klíč → ``"led"``/``"opp"``.

        Klíče: **osobní číslo** (Axxxxx, uppercase) — primární a jednoznačné, u
        vedených i **oponovaných** (``OpposingThesis.student_university_id``) —
        a jako fallback foldované **jméno bez titulů** (v rozpisu PDF mají
        studenti tituly, v práci ne, tak by se jinak nespárovali).
        """
        from .komise_stats import student_name_key

        out: dict[str, str] = {}
        for t in self.list_theses():
            student = self.get_student(t.student_id) if t.student_id else None
            if student is None:
                continue
            if student.university_id:
                out[student.university_id.strip().upper()] = "led"
            full = student_name_key(f"{student.first_name} {student.last_name}")
            if full:
                out.setdefault(full, "led")
        for o in self.list_opposing_theses():
            if o.student_university_id:
                out.setdefault(o.student_university_id.strip().upper(), "opp")
            full = student_name_key(f"{o.student_first_name} {o.student_last_name}")
            if full:
                out.setdefault(full, "opp")
        return out

    @staticmethod
    def _sched_sort_key(date: str, time: str):
        """Klíč pro chronologické řazení slotu (rok, měsíc, den, čas)."""
        m = re.search(r"(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})", date or "")
        d = (int(m.group(3)), int(m.group(2)), int(m.group(1))) if m else (9999, 99, 99)
        return (*d, time or "")

    def my_defense_schedule(self) -> list[dict]:
        """Chronologický harmonogram obhajob **mých** studentů (vedené + oponované).

        Projde sloty všech komisí a vybere ty, kde je student vedený (``led``)
        nebo oponovaný (``opp``) — spárováno přes osobní číslo / jméno
        (:meth:`komise_student_roles`). Výsledek je seřazený podle data a času;
        každá položka nese i komisi (kde) a barvu/obor. Slouží jako osobní
        rozvrh „kdy a kde mám být".
        """
        from .komise_stats import student_name_key

        roles = self.komise_student_roles()
        out: list[dict] = []
        for c in self.list_committees():
            for s in c.slots:
                pnum = (s.personal_number or "").strip().upper()
                role = roles.get(pnum) if pnum else None
                if not role:
                    role = roles.get(student_name_key(s.student_name))
                if role not in ("led", "opp"):
                    continue
                out.append({
                    "date": s.date,
                    "time": s.time,
                    "personal_number": s.personal_number,
                    "student_name": s.student_name,
                    "role": role,
                    "committee": c.display_name,
                    "color": c.color,
                    "obor": c.obor,
                    "level": c.level,
                    "academic_year": c.academic_year,
                })
        out.sort(key=lambda e: self._sched_sort_key(e["date"], e["time"]))
        return out

    def calendar_events(
        self,
        academic_year: str,
        *,
        include_led: bool = True,
        include_opp: bool = True,
        now=None,
    ) -> list[dict]:
        """Obhajoby mých studentů ve `academic_year` jako události pro kalendář.

        Vrací list dictů připravených pro :func:`ics_export.build_ics`
        (``uid``/``start``/``end``/``summary``/``location``/``description``),
        chronologicky. Délka události dle stupně: **Bc 45 min, Mgr 60 min**.
        ``include_led``/``include_opp`` filtrují vedené/oponované; ``now``
        (datetime) ořízne již proběhlé sloty — slouží pro tlačítko „Přidat do
        kalendáře" v harmonogramu (jen nadcházející ze zvoleného roku).
        """
        from datetime import timedelta

        out: list[dict] = []
        for e in self.my_defense_schedule():
            if e["academic_year"] != academic_year:
                continue
            role = e["role"]
            if role == "led" and not include_led:
                continue
            if role == "opp" and not include_opp:
                continue
            start = self._parse_slot_dt(e["date"], e["time"])
            if start is None:
                continue
            if now is not None and start < now:
                continue
            dur = 60 if (e.get("level") or "").strip().lower() == "mgr" else 45
            student = (e.get("student_name") or "").strip() or "?"
            role_cz = "vedoucí" if role == "led" else "oponent"
            emoji = "🎓" if role == "led" else "🧐"
            loc = f"Komise {e.get('color') or ''}".strip()
            if e.get("obor"):
                loc += f" ({e['obor']})"
            desc = "\n".join(
                line for line in [
                    f"Role: {role_cz}",
                    f"Osobní číslo: {e['personal_number']}"
                    if e.get("personal_number") else "",
                    f"Komise: {e['committee']}" if e.get("committee") else "",
                    f"Stupeň: {e.get('level') or '?'}",
                    f"Akademický rok: {academic_year}",
                ] if line
            )
            uid = (
                f"{e['date']}-{e['time']}-{e.get('personal_number') or ''}-{role}"
                .replace(" ", "").replace(".", "").replace(":", "")
                + "@bpdpmanager"
            )
            out.append({
                "uid": uid,
                "start": start,
                "end": start + timedelta(minutes=dur),
                "summary": f"{emoji} Obhajoba: {student}",
                "location": loc,
                "description": desc,
                "role": role,
            })
        out.sort(key=lambda x: x["start"])
        return out

    @staticmethod
    def _parse_slot_dt(date: str, time: str):
        """„15. 6. 2026" + „09:00" → datetime, nebo None když nejde rozparsovat."""
        from datetime import datetime

        md = re.search(r"(\d{1,2})\.\s*(\d{1,2})\.\s*(\d{4})", date or "")
        mt = re.search(r"(\d{1,2}):(\d{2})", time or "")
        if not (md and mt):
            return None
        try:
            return datetime(int(md.group(3)), int(md.group(2)), int(md.group(1)),
                            int(mt.group(1)), int(mt.group(2)))
        except ValueError:
            return None

    def committee_date_range(self):
        """(nejdřívější, nejpozdější) datum napříč všemi komisemi, nebo None.

        Vymezuje „období státnic" — tichá kontrola stavu obhajob běží jen
        v tomto rozmezí (jindy není potřeba).
        """
        dates = []
        for c in self.list_committees():
            for d in c.dates:
                dt = self._parse_slot_dt(d, "00:00")
                if dt is not None:
                    dates.append(dt.date())
        return (min(dates), max(dates)) if dates else None

    def in_committee_period(self, today) -> bool:
        """Je ``today`` (datum) v období státnic (rozmezí termínů komisí)?"""
        rng = self.committee_date_range()
        return bool(rng and rng[0] <= today <= rng[1])

    def upcoming_defense_reminders(self, now, within_minutes: int = 10) -> list[dict]:
        """Obhajoby mých studentů začínající **do ``within_minutes`` minut**.

        Vrací položky harmonogramu (vedené/oponované), jejichž čas spadá do
        intervalu ⟨``now``, ``now`` + ``within_minutes``⟩ — pro připomínku „za
        10 minut jdeš ke komisi". Každá položka má navíc ``minutes`` (kolik
        zbývá) a ``key`` (stabilní identifikátor slotu pro deduplikaci
        oznámení). ``now`` se předává zvenčí (kvůli testovatelnosti).
        """
        out: list[dict] = []
        for e in self.my_defense_schedule():
            dt = self._parse_slot_dt(e["date"], e["time"])
            if dt is None:
                continue
            delta = (dt - now).total_seconds()
            if 0 <= delta <= within_minutes * 60:
                out.append({
                    **e,
                    "minutes": int(delta // 60),
                    "key": f"{e['date']}|{e['time']}|{e['personal_number']}"
                           f"|{e['student_name']}|{e['role']}",
                })
        return out

    def ensure_stag_urls(self) -> int:
        """Doplní chybějící odkaz na STAG u prací se známým STAG ID.

        Odkaz je deterministicky odvozený z ``adipidno``
        (:func:`stag_api.thesis_detail_url`), takže ho lze doplnit **bez sítě**
        — i zpětně u dříve naimportovaných prací. Plní jen **prázdné** pole
        (ručně zadaný odkaz se nepřepisuje). Idempotentní; běží při startu.
        Vrací počet doplněných odkazů (vedené + oponované).
        """
        from . import stag_api

        filled = 0
        for thesis in self.list_theses():
            if thesis.adipidno and not (thesis.stag_url or "").strip():
                thesis.stag_url = stag_api.thesis_detail_url(thesis.adipidno)
                self.upsert_thesis(thesis)
                filled += 1
        for op in self.list_opposing_theses():
            if op.adipidno and not (op.stag_url or "").strip():
                op.stag_url = stag_api.thesis_detail_url(op.adipidno)
                self.upsert_opposing_thesis(op)
                filled += 1
        return filled

    def repair_review_archive_names(self) -> int:
        """Jednorázová oprava názvů archivních posudků.

        Sloučí vícenásobně zanořené ``_archiv_<ts>`` segmenty v názvech
        archivovaných souborů (důsledek staršího bugu, kdy se archiv připisoval
        i už archivovaným souborům). Idempotentní — čisté názvy nechá být.
        Vrací počet přejmenovaných souborů.
        """
        repaired = 0
        for thesis in self.list_theses():
            n = self._repair_archive_names_for(thesis.id, thesis.attachments)
            if n:
                self.upsert_thesis(thesis)
                repaired += n
        for op in self.list_opposing_theses():
            n = self._repair_archive_names_for(f"opposing-{op.id}", op.attachments)
            if n:
                self.upsert_opposing_thesis(op)
                repaired += n
        return repaired

    def _repair_archive_names_for(
        self, container_id: str, attachments: list[Attachment]
    ) -> int:
        base = thesis_documents_dir(container_id)
        renamed = 0
        for att in attachments:
            if not att.is_file:
                continue
            rel = Path(att.url_or_path)
            new_name = self._NESTED_ARCHIVE_RE.sub(r"\1", rel.name)
            if new_name == rel.name:
                continue
            new_rel = rel.with_name(new_name)
            old_abs = base / rel
            new_abs = base / new_rel
            # Zajisti unikátnost, kdyby cílový název už existoval.
            if new_abs != old_abs and new_abs.exists():
                stem, suffix = new_rel.stem, new_rel.suffix
                k = 2
                while (base / new_rel.with_name(f"{stem}_{k}{suffix}")).exists():
                    k += 1
                new_rel = new_rel.with_name(f"{stem}_{k}{suffix}")
                new_abs = base / new_rel
            try:
                if old_abs.exists():
                    old_abs.rename(new_abs)
                old_basename = rel.name
                att.url_or_path = str(new_rel)
                if att.label == old_basename:
                    att.label = new_rel.name
                renamed += 1
            except OSError:
                pass
        return renamed

    def prune_missing_documents(self, thesis_id: str) -> int:
        """Odebere ze seznamu příloh ty soubory, které fyzicky neexistují.

        Vrací počet odebraných záznamů. URL odkazy a existující soubory zůstávají.
        """
        thesis = self.get_thesis(thesis_id)
        if thesis is None:
            return 0
        base_dir = thesis_documents_dir(thesis_id)
        kept = [
            att for att in thesis.attachments
            if not att.is_file or (base_dir / att.url_or_path).exists()
        ]
        removed = len(thesis.attachments) - len(kept)
        if removed:
            thesis.attachments = kept
            self.upsert_thesis(thesis)
        return removed

    def opposing_prune_missing_documents(self, op_id: str) -> int:
        """Obdoba ``prune_missing_documents`` pro oponentský posudek."""
        op = self.get_opposing_thesis(op_id)
        if op is None:
            return 0
        base_dir = thesis_documents_dir(f"opposing-{op_id}")
        kept = [
            att for att in op.attachments
            if not att.is_file or (base_dir / att.url_or_path).exists()
        ]
        removed = len(op.attachments) - len(kept)
        if removed:
            op.attachments = kept
            self.upsert_opposing_thesis(op)
        return removed

    @staticmethod
    def _template_cell_text(tmpl_path: Path, cell: str | None) -> str | None:
        """Přečte (read-only) text buňky šablony — pro place_date transformaci."""
        if not cell:
            return None
        try:
            wb = load_template_workbook(tmpl_path, data_only=False)
            ws = wb.active
            val = ws[cell].value
            wb.close()
            if val is None:
                return None
            return val if isinstance(val, str) else str(val)
        except Exception:  # noqa: BLE001
            return None

    @staticmethod
    def _estimate_text_row_heights(
        tmpl_path: Path, text_cells: dict[str, str]
    ) -> dict[int, float]:
        """Odhad výšky řádků pro buňky s dlouhým textem (aby se PDF neuseklo).

        ``text_cells``: ``{souřadnice: text}``. Vrací ``{řádek: výška_v_bodech}``.
        Best-effort — při chybě (chybí openpyxl, neznámá geometrie) vrací {}.
        Bere v úvahu sloučené buňky (sečte šířky sloupců) a explicitní zalomení.
        """
        if not text_cells:
            return {}
        try:
            from openpyxl.utils.cell import (
                column_index_from_string,
                coordinate_from_string,
                get_column_letter,
            )
            wb = load_template_workbook(tmpl_path, data_only=False)
            ws = wb.active
        except Exception:
            return {}
        try:
            merges = list(ws.merged_cells.ranges)

            def span(coord: str) -> tuple[int, int, int]:
                letter, row = coordinate_from_string(coord)
                col = column_index_from_string(letter)
                for rng in merges:
                    mc, mr, xc, xr = rng.bounds
                    if mc <= col <= xc and mr <= row <= xr:
                        return mc, xc, row
                return col, col, row

            heights: dict[int, float] = {}
            for coord, text in text_cells.items():
                if not text:
                    continue
                min_col, max_col, row = span(coord)
                total_w = 0.0
                for ci in range(min_col, max_col + 1):
                    dim = ws.column_dimensions.get(get_column_letter(ci))
                    total_w += dim.width if dim and dim.width else 8.43
                # Excelová šířka ≈ počet znaků; mírně podhodnotíme (→ vyšší řádek,
                # raději trochu místa navíc než useknutý text).
                chars_per_line = max(12.0, total_w * 0.95)
                lines = 0
                for line in str(text).split("\n"):
                    lines += max(1, math.ceil(len(line) / chars_per_line))
                if lines < 3:
                    continue  # krátký text — výšku necháme na šabloně
                heights[row] = max(
                    heights.get(row, 0.0), min(lines * 15.0 + 6.0, 2000.0)
                )
            return heights
        except Exception:
            return {}
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # Kolik řádků textu se rozumně vejde na jednu stránku posudku (heuristika).
    _PAGE_LINES = 40
    _LINE_PT = 15.0

    @staticmethod
    def _cell_style_index(tmpl_path: Path, coord: str) -> str | None:
        """Přečte index stylu (``s``) buňky z XML šablony (pro kopii stylu)."""
        import zipfile

        from .xlsx_cell_writer import _resolve_active_sheet_part
        try:
            with zipfile.ZipFile(tmpl_path) as zf:
                xml = zf.read(_resolve_active_sheet_part(zf)).decode("utf-8")
            m = re.search(rf'<c\s+r="{re.escape(coord)}"[^>]*?\bs="(\d+)"', xml)
            return m.group(1) if m else None
        except Exception:
            return None

    @staticmethod
    def _split_text_into_parts(
        text: str, chars_per_line: float, max_parts: int, page_lines: int
    ) -> list[str]:
        """Rozdělí text na ≤ ``max_parts`` částí na hranicích odstavců tak, aby
        každá (kromě poslední) měla zhruba ``page_lines`` řádků."""
        def plines(s: str) -> int:
            return max(1, math.ceil(len(s) / chars_per_line))

        parts: list[str] = []
        cur: list[str] = []
        cur_lines = 0
        for para in text.split("\n"):
            pl = plines(para)
            if (cur and cur_lines + pl > page_lines
                    and len(parts) < max_parts - 1):
                parts.append("\n".join(cur))
                cur, cur_lines = [], 0
            cur.append(para)
            cur_lines += pl
        if cur:
            parts.append("\n".join(cur))
        return parts

    def _plan_long_text(
        self, tmpl_path: Path, cell: str, text: str
    ) -> tuple[dict[str, str], dict[str, list[str]], dict[int, float],
               dict[str, str]]:
        """Naplánuje zápis dlouhého textu do (sloučené) buňky.

        Krátký text → jedna buňka. Dlouhý text, který se nevejde na stránku, se
        rozdělí přes řádky sloučené buňky (rozbití slučení + kopie stylu), aby
        v PDF tekl přes stránky. Vrací ``(values, merge_splits, row_heights,
        cell_styles)``.
        """
        single = ({cell: text}, {}, {}, {})
        if not text:
            return single
        try:
            from openpyxl.utils.cell import (
                column_index_from_string,
                coordinate_from_string,
                get_column_letter,
            )
            wb = load_template_workbook(tmpl_path, data_only=False)
            ws = wb.active
        except Exception:
            return single
        try:
            col_letter, row = coordinate_from_string(cell)
            col = column_index_from_string(col_letter)
            mc = xc = col
            mr = xr = row
            for rng in ws.merged_cells.ranges:
                bmc, bmr, bxc, bxr = rng.bounds
                if bmc <= col <= bxc and bmr <= row <= bxr:
                    mc, mr, xc, xr = bmc, bmr, bxc, bxr
                    break
            total_w = 0.0
            for ci in range(mc, xc + 1):
                dim = ws.column_dimensions.get(get_column_letter(ci))
                total_w += dim.width if dim and dim.width else 8.43
            cpl = max(12.0, total_w * 0.95)
            total_lines = sum(
                max(1, math.ceil(len(line) / cpl)) for line in text.split("\n")
            )
            nrows = xr - mr + 1

            if total_lines < 3:
                return single  # krátké — výšku necháme na šabloně
            # Vejde se na stránku (nebo nelze rozdělit) → jedna buňka + výška.
            if total_lines <= self._PAGE_LINES or nrows < 2:
                h = min(total_lines * self._LINE_PT + 6.0, 2000.0)
                return {cell: text}, {}, {row: h}, {}

            parts = self._split_text_into_parts(text, cpl, nrows, self._PAGE_LINES)
            if len(parts) < 2:
                # Jeden velký odstavec → rozdělit nešlo; aspoň vyšší buňka.
                h = min(total_lines * self._LINE_PT + 6.0, 2000.0)
                return {cell: text}, {}, {row: h}, {}

            left = get_column_letter(mc)
            right = get_column_letter(xc)
            old_ref = f"{left}{mr}:{right}{xr}"
            new_refs = [f"{left}{mr + i}:{right}{mr + i}" for i in range(nrows)]
            style = self._cell_style_index(tmpl_path, cell)
            values: dict[str, str] = {}
            heights: dict[int, float] = {}
            styles: dict[str, str] = {}
            for i, part in enumerate(parts):
                r = mr + i
                coord = f"{left}{r}"
                values[coord] = part
                lines = sum(
                    max(1, math.ceil(len(line) / cpl))
                    for line in part.split("\n")
                )
                heights[r] = min(lines * self._LINE_PT + 6.0, 2000.0)
                if i > 0 and style:  # druhá+ buňka potřebuje styl té první
                    styles[coord] = style
            return values, {old_ref: new_refs}, heights, styles
        except Exception:
            return single
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _xlsx_to_pdf(self, xlsx_path: Path) -> Path | None:
        """Konvertuje XLSX na PDF přes LibreOffice headless.

        Vrátí cestu k vytvořenému PDF, nebo ``None`` pokud LibreOffice
        není dostupný / konverze selhala. Volající by měl gracefully
        pokračovat bez PDF.

        Před převodem se na **dočasné kopii** zploští případný „obrázek
        v buňce" (Excel rich-value image) na klasický plovoucí obrázek —
        LibreOffice ho jinak neumí vykreslit a v PDF chybí logo + zobrazí
        ``#VALUE!``. Uložený XLSX zůstává s nativním obrázkem v buňce
        (Excel uživatel ho má 1:1).
        """
        import subprocess
        import tempfile

        from .xlsx_image_in_cell import flatten_image_in_cell
        from .xlsx_pdf_polish import polish_pdf_layout

        soffice = self._find_soffice()
        if soffice is None:
            return None

        pdf_path = xlsx_path.with_suffix(".pdf")
        with tempfile.TemporaryDirectory() as tmp:
            tmp_dir = Path(tmp)
            convert_src = tmp_dir / xlsx_path.name
            try:
                flatten_image_in_cell(xlsx_path, convert_src)
                # Kosmetika jen pro PDF: vyvážené okraje + menší/černá hlavička
                # „Body". Pracuje na dočasné kopii, uložený XLSX zůstává 1:1.
                polish_pdf_layout(convert_src)
            except Exception:  # noqa: BLE001 — fail-safe, převeď aspoň originál
                convert_src = xlsx_path
            # Izolovaný uživatelský profil LibreOffice — KRITICKÉ: bez něj
            # ``soffice --headless`` zamrzne (čeká na zámek profilu), když má
            # uživatel LibreOffice otevřené v GUI. Vlastní profil v tempu
            # konflikt obejde a konverzi nikdy neblokuje běžící instance.
            lo_profile = (tmp_dir / "loprofile").as_uri()
            try:
                proc = subprocess.run(
                    [
                        str(soffice),
                        f"-env:UserInstallation={lo_profile}",
                        "--headless",
                        "--convert-to", "pdf",
                        "--outdir", str(tmp_dir),
                        str(convert_src),
                    ],
                    capture_output=True,
                    timeout=120,
                    check=False,
                )
            except (subprocess.TimeoutExpired, OSError):
                return None
            produced = tmp_dir / (convert_src.stem + ".pdf")
            if proc.returncode != 0 or not produced.is_file():
                return None
            try:
                shutil.move(str(produced), str(pdf_path))
            except OSError:
                return None
        return pdf_path

    @staticmethod
    def _find_soffice() -> Path | None:
        """Najde ``soffice`` binární soubor LibreOffice."""
        import shutil as _shutil

        # 1) Standardní PATH
        bin_path = _shutil.which("soffice") or _shutil.which("libreoffice")
        if bin_path:
            return Path(bin_path)

        # 2) macOS standardní lokace
        mac_path = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
        if mac_path.is_file():
            return mac_path

        # 3) Linux distro typicky /usr/bin/soffice
        for candidate in ("/usr/bin/soffice", "/usr/local/bin/soffice"):
            p = Path(candidate)
            if p.is_file():
                return p

        return None

    @property
    def libreoffice_available(self) -> bool:
        """True pokud je LibreOffice nainstalován a dostupný pro PDF export."""
        return self._find_soffice() is not None
