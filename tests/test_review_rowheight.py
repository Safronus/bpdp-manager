"""Výška řádku pro dlouhý text v posudku — ať se v PDF (LibreOffice) neusekne."""

from __future__ import annotations

from pathlib import Path

from bpdpmanager.services.thesis_service import ThesisService
from bpdpmanager.services.xlsx_cell_writer import set_cells


def _make_xlsx(path: Path, *, cell: str = "B2", value: str = "x",
               merge: str | None = None, widths: dict | None = None) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws[cell] = value
    if merge:
        ws.merge_cells(merge)
    for col, w in (widths or {}).items():
        ws.column_dimensions[col].width = w
    wb.save(path)


def test_set_cells_applies_row_height(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    tpl = tmp_path / "t.xlsx"
    _make_xlsx(tpl, cell="A1", value="orig")
    out = tmp_path / "o.xlsx"
    set_cells(tpl, out, {"A1": "nový"}, row_heights={1: 180.0})
    ws = load_workbook(out).active
    assert ws["A1"].value == "nový"
    assert ws.row_dimensions[1].height == 180.0
    assert ws.row_dimensions[1].customHeight


def test_long_text_gets_tall_row(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_xlsx(tpl, merge="B2:E2",
               widths={"B": 20, "C": 20, "D": 20, "E": 20})
    heights = ThesisService._estimate_text_row_heights(tpl, {"B2": "x" * 600})
    assert 2 in heights
    assert heights[2] > 90  # ~8 radku po 15 bodech


def test_short_text_keeps_template_height(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_xlsx(tpl, merge="B2:E2",
               widths={"B": 20, "C": 20, "D": 20, "E": 20})
    assert ThesisService._estimate_text_row_heights(tpl, {"B2": "krátký"}) == {}


def test_explicit_newlines_count_as_lines(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_xlsx(tpl, merge="B2:E2",
               widths={"B": 20, "C": 20, "D": 20, "E": 20})
    # pět krátkých řádků díky \n → ≥ 3 řádky → výška se nastaví
    heights = ThesisService._estimate_text_row_heights(
        tpl, {"B2": "a\nb\nc\nd\ne"}
    )
    assert 2 in heights


def test_empty_text_cells_returns_empty(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_xlsx(tpl)
    assert ThesisService._estimate_text_row_heights(tpl, {}) == {}


# ── rozdělení dlouhého textu do dvou polí (jen když nutné) ──────────────────

def _make_merged(path: Path, ref: str = "B2:E3",
                 widths: dict | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    wb = Workbook()
    ws = wb.active
    ws["B2"] = "x"
    ws["B2"].alignment = Alignment(wrap_text=True)  # → B2 dostane styl s wrapem
    ws.merge_cells(ref)
    for col, w in (widths or {"B": 20, "C": 20, "D": 20, "E": 20}).items():
        ws.column_dimensions[col].width = w
    wb.save(path)


def _svc() -> ThesisService:
    return ThesisService.__new__(ThesisService)


def test_long_text_splits_into_two_fields(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_merged(tpl, "B2:E3")
    vals, splits, heights, styles = _svc()._plan_long_text(tpl, "B2", "Odstavec.\n" * 60)
    assert set(vals) == {"B2", "B3"}              # dvě pole
    assert splits == {"B2:E3": ["B2:E2", "B3:E3"]}  # sloučení rozbito po řádcích
    assert "B3" in styles                          # druhá buňka má styl té první
    assert 2 in heights and 3 in heights


def test_short_text_stays_single_field(tmp_path: Path) -> None:
    tpl = tmp_path / "t.xlsx"
    _make_merged(tpl, "B2:E3")
    vals, splits, _heights, styles = _svc()._plan_long_text(tpl, "B2", "krátké hodnocení")
    assert list(vals) == ["B2"]
    assert splits == {} and styles == {}


def test_replace_merges_updates_count() -> None:
    from bpdpmanager.services.xlsx_cell_writer import _replace_merges

    xml = ('<mergeCells count="2"><mergeCell ref="A1:B1"/>'
           '<mergeCell ref="A43:D44"/></mergeCells>')
    out = _replace_merges(xml, {"A43:D44": ["A43:D43", "A44:D44"]})
    assert 'count="3"' in out
    assert 'ref="A43:D43"' in out and 'ref="A44:D44"' in out
    assert 'ref="A43:D44"' not in out


def test_split_copies_style_to_second_cell(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    from bpdpmanager.services.xlsx_cell_writer import set_cells

    tpl = tmp_path / "t.xlsx"
    _make_merged(tpl, "B2:E3")
    vals, splits, heights, styles = _svc()._plan_long_text(tpl, "B2", "Odstavec.\n" * 60)
    out = tmp_path / "o.xlsx"
    set_cells(tpl, out, vals, row_heights=heights, merge_splits=splits,
              cell_styles=styles)
    ws = load_workbook(out).active
    # druhá buňka má stejný styl jako první (wrap) → text se v PDF neořízne
    assert ws["B2"].style == ws["B3"].style
    assert ws["B3"].alignment.wrap_text
