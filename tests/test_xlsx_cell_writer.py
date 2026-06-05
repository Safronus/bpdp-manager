"""Testy XML-level zapisovače buněk XLSX.

Klíčový požadavek: vyplnění posudku musí být **1:1 se šablonou** — zejména
nesmí zmizet logo (obrázek). Tyto testy ověřují, že:

- vložený obrázek (xl/media + xl/drawings) přežije zápis hodnot,
- string i číselné hodnoty se zapíšou správně,
- chybějící buňky/řádky se vytvoří,
- styl buňky (`s`) zůstane zachován.
"""

from __future__ import annotations

import zipfile
from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
PILImage = pytest.importorskip("PIL.Image")

from bpdpmanager.services import xlsx_cell_writer as xw  # noqa: E402


def _media_parts(path: Path) -> list[str]:
    with zipfile.ZipFile(path) as z:
        return sorted(n for n in z.namelist() if "media" in n or "drawing" in n)


@pytest.fixture
def logo_template(tmp_path: Path) -> Path:
    """XLSX šablona s logem + popisky + (stylovanou) prázdnou buňkou."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    png = tmp_path / "logo.png"
    PILImage.new("RGB", (60, 30), (0, 90, 160)).save(str(png))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A6"] = "Student:"
    ws["B6"] = "PŮVODNÍ"
    ws["A18"] = "Kritérium"
    ws["D18"].font = Font(bold=True)  # stylovaná prázdná buňka
    ws["A49"] = "Místo, datum: ........  Podpis: ........"
    ws.add_image(XLImage(str(png)), "F1")
    path = tmp_path / "tmpl.xlsx"
    wb.save(str(path))
    return path


def test_logo_preserved(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"B6": "Jan Novák"})
    assert _media_parts(out) == _media_parts(logo_template)
    assert _media_parts(out)  # opravdu tam logo je


def test_string_value_overwrite(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"B6": "Jan Novák"})
    wb = openpyxl.load_workbook(out)
    assert wb.active["B6"].value == "Jan Novák"
    assert wb.active["A6"].value == "Student:"  # label nedotčen


def test_number_into_styled_empty_cell_keeps_style(
    logo_template: Path, tmp_path: Path
) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"D18": 4})
    wb = openpyxl.load_workbook(out)
    cell = wb.active["D18"]
    assert cell.value == 4
    assert cell.font.bold is True  # styl zachován


def test_creates_missing_cell_and_row(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"C18": 10.5, "B40": "Nový řádek"})
    wb = openpyxl.load_workbook(out)
    assert wb.active["C18"].value == 10.5
    assert wb.active["B40"].value == "Nový řádek"


def test_float_integer_written_as_int(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"C18": 5.0})
    wb = openpyxl.load_workbook(out)
    # 5.0 → uloženo jako 5 (číslo), openpyxl načte int/float 5
    assert wb.active["C18"].value == 5


def test_place_date_partial_replace(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    # Simuluje transformaci „Místo, datum: …" — nahrazení první tečkované linky
    import re

    existing = "Místo, datum: ........  Podpis: ........"
    new_text = re.sub(r"\.{3,}", " Zlín 5.6.2026 ", existing, count=1)
    xw.set_cells(logo_template, out, {"A49": new_text})
    wb = openpyxl.load_workbook(out)
    val = wb.active["A49"].value
    assert "Zlín 5.6.2026" in val
    assert "Podpis: ........" in val  # podpisový blok zachován


def test_empty_values_skipped(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"B6": "", "A6": None})
    wb = openpyxl.load_workbook(out)
    # Prázdné/None hodnoty se nezapisují — originál zůstane
    assert wb.active["B6"].value == "PŮVODNÍ"
    assert wb.active["A6"].value == "Student:"


def test_invalid_coord_raises(logo_template: Path, tmp_path: Path) -> None:
    out = tmp_path / "out.xlsx"
    with pytest.raises(xw.XlsxWriteError):
        xw.set_cells(logo_template, out, {"NOTACOORD": "x"})


def test_col_to_index() -> None:
    assert xw._col_to_index("A") == 1
    assert xw._col_to_index("Z") == 26
    assert xw._col_to_index("AA") == 27
    assert xw._col_to_index("AB") == 28


def test_clear_formula_caches_strips_value_keeps_formula() -> None:
    # Numerická i řetězcová cache + self-closing varianta
    xml = (
        '<c r="B30"><f>C18*D18</f><v>0</v></c>'
        '<c r="B32"><f>IF(B31&gt;=50,"E","F")</f><v>F</v></c>'
        '<c r="B33"><f>SUM(A1:A2)</f><v/></c>'
    )
    out = xw._clear_formula_caches(xml)
    assert "<f>C18*D18</f>" in out
    assert "<v>0</v>" not in out
    assert "<v>F</v>" not in out
    # formula bez cache zůstane formulí
    assert out.count("<f>") == 3


def test_clear_formula_caches_leaves_plain_values() -> None:
    # Buňka bez <f> (prostá hodnota) se nesmí dotknout
    xml = '<c r="A1" t="inlineStr"><is><t>Label</t></is></c><c r="A2"><v>5</v></c>'
    assert xw._clear_formula_caches(xml) == xml


def test_force_full_calc_adds_attribute() -> None:
    wb = '<workbook><sheets/><calcPr calcId="124519"/></workbook>'
    out = xw._force_full_calc(wb)
    assert 'fullCalcOnLoad="1"' in out


def test_force_full_calc_replaces_zero() -> None:
    wb = '<workbook><calcPr calcId="1" fullCalcOnLoad="0"/></workbook>'
    out = xw._force_full_calc(wb)
    assert 'fullCalcOnLoad="1"' in out
    assert 'fullCalcOnLoad="0"' not in out


def test_force_full_calc_inserts_when_missing() -> None:
    wb = "<workbook><sheets/></workbook>"
    out = xw._force_full_calc(wb)
    assert "<calcPr" in out and 'fullCalcOnLoad="1"' in out


def test_set_cells_sets_full_calc_on_output(
    logo_template: Path, tmp_path: Path
) -> None:
    import zipfile

    out = tmp_path / "out.xlsx"
    xw.set_cells(logo_template, out, {"B6": "Jan"})
    with zipfile.ZipFile(out) as z:
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
    assert 'fullCalcOnLoad="1"' in wbxml
